"""
Migração: CONTRATOS_V2_2024.xlsx → contratos / contrato_parcelas (SQLite)

Uso:
    python migrate_contratos.py
    python migrate_contratos.py --force   # limpa e re-migra
"""
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

XLSX = Path("/Users/michelletadra/Downloads/CONTRATOS_V2_2024.xlsx")
FORCE = "--force" in sys.argv

# ── Mapeamento de abas ────────────────────────────────────────────────────────
# (nome_aba, status_override)  — None = detecta pela célula SITUAÇÃO
SHEETS = [
    # Ativos
    ("AGROCETE",       None),
    ("NITRO1000 2026", None),
    ("TRADECORP 2026", None),
    ("GGSCOOPACER",    None),
    ("SATIS",          None),
    ("GSCOOPACER",     None),
    ("SOLUBIO",        None),
    # Encerrados
    ("TRADECORP 2025",        "ENCERRADO"),
    ("NITRO1000 2025",        "ENCERRADO"),
    ("TRADECORP 2024",        "ENCERRADO"),
    ("GGCOOPACER",            "ENCERRADO"),
    ("GGSCOOPACER 2024",      "ENCERRADO"),
    ("NITRO 1000 ENCERRADO",  "ENCERRADO"),
    ("GS COOPACER ENCERRADO", "ENCERRADO"),
    ("TECSOLO",               "ENCERRADO"),
    ("GREENBACK",             "ENCERRADO"),
    ("GPXTECNOLOGIA",         "ENCERRADO"),
]

def to_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date().isoformat()
    s = str(val).strip()
    if s.startswith("20") and len(s) == 10:
        return s
    return None

def _safe_nf(val):
    if not val:
        return None
    s = str(val).strip()
    if s in ("", "None", "MÁXIMO", "ENTREGUE", "FALTA"):
        return None
    try:
        return str(int(float(s)))
    except Exception:
        return None

def to_float(val):
    if val is None:
        return 0.0
    try:
        return float(val)
    except Exception:
        return 0.0

def parse_sheet(ws):
    """Extrai dados de contrato e parcelas de uma aba."""
    rows = list(ws.iter_rows(values_only=True))

    # Mapa vertical: col[1] (label) → col[2] (valor) para linhas 5-20
    spec = {}
    for row in rows[4:20]:
        if row[1] and row[2] is not None:
            spec[str(row[1]).strip().upper()] = row[2]

    contrato = {
        "empresa":         str(spec.get("EMPRESA", "") or "").strip(),
        "contratante":     str(spec.get("CONTRATANTE", "") or "").strip(),
        "data_assinatura": to_date(spec.get("ASSINATURA") or spec.get("RENOVAÇÃO") or spec.get("RENOVAÇÃO AUTOMÁTICA")),
        "data_vencimento": to_date(spec.get("VALIDADE")),
        "valor_total":     to_float(spec.get("CONTRATO")),
        "valor_parcela":   to_float(spec.get("PARCELA")),
        "valor_consumido": to_float(spec.get("CONSUMO")),
        "saldo":           to_float(spec.get("SALDO")),
        "situacao":        str(spec.get("SITUAÇÃO", "") or "").strip(),
        "servico":         str(spec.get("SERVIÇO", "") or spec.get("SERVIÇOS", "") or "").strip() or None,
        "kit_info":        str(spec.get("KIT", "") or "").strip() or None,
        "amostras_info":   str(spec.get("AMOSTRAS", "") or "").strip() or None,
        "info_pagamento":  str(spec.get("INFORMAÇÕES", "") or "").strip() or None,
        "observacoes":     str(spec.get("OBS", "") or "").strip() or None,
    }

    # Parcelas: header em row[4] (índice), dados a partir de row[5]
    # Colunas típicas: PARCELA(4), EMISSÃO(5), VALOR(6), SALDO ATUAL(7), SITUAÇÃO(8), NF(9)
    # Detecta offset da coluna de parcelas pelo header
    header_row = rows[4]  # row 5 (0-indexed = 4)
    parc_col = None
    for ci, cell in enumerate(header_row):
        if str(cell or "").strip().upper() == "PARCELA":
            parc_col = ci
            break

    parcelas = []
    if parc_col is not None:
        for row in rows[5:]:
            num = row[parc_col]
            if num is None:
                continue
            try:
                num_int = int(float(str(num)))
            except Exception:
                continue
            if num_int <= 0:
                continue
            parcelas.append({
                "numero":       num_int,
                "data_emissao": to_date(row[parc_col + 1]),
                "valor":        to_float(row[parc_col + 2]),
                "saldo_atual":  to_float(row[parc_col + 3]),
                "situacao":     str(row[parc_col + 4] or "").strip() or None,
                "numero_nf":    _safe_nf(row[parc_col + 5]),
            })

    contrato["total_parcelas"] = len(parcelas)
    return contrato, parcelas

# ── Importa módulo de banco ───────────────────────────────────────────────────
import sys, os
sys.path.insert(0, str(Path(__file__).parent))
from db_contratos_sqlite import (
    init_contratos, insert_contrato, insert_parcela, list_contratos, _conn
)

init_contratos()

# Verifica se já tem dados
existing = _conn().execute("SELECT COUNT(*) FROM contratos").fetchone()[0]
if existing > 0:
    if FORCE:
        with _conn() as conn:
            conn.execute("DELETE FROM contrato_parcelas")
            conn.execute("DELETE FROM contratos")
        print(f"⚠️  --force: {existing} contrato(s) removido(s)")
    else:
        print(f"⚠️  Já existem {existing} contrato(s). Use --force para re-migrar.")
        sys.exit(1)

# ── Migração ──────────────────────────────────────────────────────────────────
print(f"📂 Lendo {XLSX.name} …")
wb = openpyxl.load_workbook(str(XLSX), read_only=True, data_only=True)

ok = 0
skipped = []
total_parcelas = 0

for sheet_name, status_override in SHEETS:
    if sheet_name not in wb.sheetnames:
        skipped.append(sheet_name)
        continue

    ws = wb[sheet_name]
    contrato, parcelas = parse_sheet(ws)

    if not contrato["contratante"]:
        skipped.append(sheet_name)
        continue

    # Status
    if status_override:
        contrato["status"] = status_override
    else:
        s = contrato["situacao"].upper()
        contrato["status"] = "ENCERRADO" if "ENCERRADO" in s else "ATIVO"

    contrato["nome_aba"] = sheet_name

    cid = insert_contrato(contrato)
    for p in parcelas:
        p["contrato_id"] = cid
        insert_parcela(p)

    total_parcelas += len(parcelas)
    ok += 1
    status_icon = "🟢" if contrato["status"] == "ATIVO" else "⚪"
    print(f"  {status_icon} {sheet_name:<25} {contrato['empresa']:<15} {contrato['contratante']:<12} "
          f"{len(parcelas)} parcelas")

wb.close()

print()
if skipped:
    print(f"  ⏭️  Abas ignoradas: {', '.join(skipped)}")
print(f"\n🎉 {ok} contratos e {total_parcelas} parcelas migrados!")
