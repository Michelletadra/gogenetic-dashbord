"""
Seed completo de contratos: planilha Excel operacional + base estratégica.

Uso:
    python seed_contratos.py           # aborta se já tiver dados
    python seed_contratos.py --force   # limpa e re-seed
"""
import sys
from datetime import datetime
from pathlib import Path

FORCE = "--force" in sys.argv

sys.path.insert(0, str(Path(__file__).parent))
from db_contratos_sqlite import _conn, init_contratos, insert_contrato, insert_parcela

init_contratos()

existing = _conn().execute("SELECT COUNT(*) FROM contratos").fetchone()[0]
if existing > 0:
    if FORCE:
        with _conn() as c:
            c.execute("DELETE FROM contrato_aditivos")
            c.execute("DELETE FROM contrato_parcelas")
            c.execute("DELETE FROM contratos")
        print(f"⚠️  {existing} contratos removidos. Re-seeding...")
    else:
        print(f"⚠️  Já existem {existing} contratos. Use --force para re-seed.")
        sys.exit(1)

# ══════════════════════════════════════════════════════════════════════════════
# PARTE 1 — Contratos operacionais da planilha Excel (com parcelas)
# ══════════════════════════════════════════════════════════════════════════════
import openpyxl

XLSX = Path("/Users/michelletadra/Downloads/CONTRATOS_V2_2024.xlsx")

SHEETS_CONFIG = [
    # (nome_aba, empresa_gg, tipo_contrato, categoria, renovacao_auto, recorrente, internacional, status_override)
    ("AGROCETE",       "GG SOLUÇÕES", "Prestação laboratorial", "Kits moleculares",  False, True,  False, None),
    ("NITRO1000 2026", "GG SOLUÇÕES", "Consultoria + kits",     "Kits moleculares",  False, True,  False, None),
    ("TRADECORP 2026", "GG SOLUÇÕES", "Comodato",               "Kits moleculares",  True,  True,  False, None),
    ("GGSCOOPACER",    "GG SOLUÇÕES", "Prestação laboratorial", "Kits moleculares",  True,  True,  False, None),
    ("SATIS",          "GG SOLUÇÕES", "Prestação laboratorial", "Kits moleculares",  True,  True,  False, None),
    ("GSCOOPACER",     "GOSOLOS",     "Prestação laboratorial", "Microbioma de solo",True,  True,  False, None),
    ("SOLUBIO",        "GOSOLOS",     "Prestação laboratorial", "Microbioma de solo",False, True,  False, None),
    # Encerrados
    ("TRADECORP 2025",        "GG SOLUÇÕES","Prestação laboratorial","Kits moleculares",False,True,False,"ENCERRADO"),
    ("NITRO1000 2025",        "GG SOLUÇÕES","Consultoria + kits","Kits moleculares",False,True,False,"ENCERRADO"),
    ("TRADECORP 2024",        "GG SOLUÇÕES","Prestação laboratorial","Kits moleculares",False,True,False,"ENCERRADO"),
    ("GGCOOPACER",            "GOGENETIC","Prestação laboratorial","Kits moleculares",False,True,False,"ENCERRADO"),
    ("GGSCOOPACER 2024",      "GG SOLUÇÕES","Prestação laboratorial","Kits moleculares",False,True,False,"ENCERRADO"),
    ("NITRO 1000 ENCERRADO",  "GG SOLUÇÕES","Consultoria + kits","Kits moleculares",False,True,False,"ENCERRADO"),
    ("GS COOPACER ENCERRADO", "GOSOLOS","Prestação laboratorial","Microbioma de solo",False,True,False,"ENCERRADO"),
    ("TECSOLO",               "GOSOLOS","Representação comercial","Canal/parceria",False,False,False,"ENCERRADO"),
    ("GREENBACK",             "GOSOLOS","Prestação de serviço","Laboratorial",False,False,False,"ENCERRADO"),
    ("GPXTECNOLOGIA",         "GOSOLOS","Prestação de serviço","Laboratorial",False,False,False,"ENCERRADO"),
]

def _to_date(val):
    if val is None: return None
    if isinstance(val, datetime): return val.date().isoformat()
    s = str(val).strip()
    return s if (s.startswith("20") and len(s) == 10) else None

def _to_float(val):
    if val is None: return 0.0
    try: return float(val)
    except: return 0.0

def _safe_nf(val):
    if not val: return None
    s = str(val).strip()
    if s in ("","None","MÁXIMO","ENTREGUE","FALTA"): return None
    try: return str(int(float(s)))
    except: return None

print("📂 Importando planilha Excel...")
wb = openpyxl.load_workbook(str(XLSX), read_only=True, data_only=True)

for (nome_aba, empresa_gg, tipo, categoria, renov, recorr, intern, st_override) in SHEETS_CONFIG:
    if nome_aba not in wb.sheetnames:
        print(f"  ⚠️  Aba '{nome_aba}' não encontrada")
        continue
    ws = wb[nome_aba]
    rows = list(ws.iter_rows(values_only=True))
    spec = {}
    for row in rows[4:20]:
        if row[1] and row[2] is not None:
            spec[str(row[1]).strip().upper()] = row[2]

    situacao = str(spec.get("SITUAÇÃO","") or "").strip()
    status   = st_override if st_override else ("ENCERRADO" if "ENCERRADO" in situacao.upper() else "ATIVO")

    contrato = {
        "nome_aba":           nome_aba,
        "empresa_gg":         empresa_gg,
        "contratante":        str(spec.get("CONTRATANTE","") or nome_aba).strip(),
        "tipo_contrato":      tipo,
        "categoria":          categoria,
        "data_assinatura":    _to_date(spec.get("ASSINATURA") or spec.get("RENOVAÇÃO")),
        "data_termino":       _to_date(spec.get("VALIDADE")),
        "valor_total":        _to_float(spec.get("CONTRATO")),
        "valor_parcela":      _to_float(spec.get("PARCELA")),
        "valor_recorrente":   _to_float(spec.get("PARCELA")),
        "amostras_contratadas": 0,
        "status":             status,
        "situacao":           situacao or None,
        "servico_principal":  str(spec.get("SERVIÇO","") or spec.get("SERVIÇOS","") or "").strip() or None,
        "kit_info":           str(spec.get("KIT","") or "").strip() or None,
        "obs_tecnicas":       str(spec.get("AMOSTRAS","") or "").strip() or None,
        "info_pagamento":     str(spec.get("INFORMAÇÕES","") or "").strip() or None,
        "observacoes":        str(spec.get("OBS","") or "").strip() or None,
        "renovacao_automatica": 1 if renov else 0,
        "recorrente":           1 if recorr else 0,
        "internacional":        1 if intern else 0,
    }

    # Remove campo não mapeado
    for k in ["kit_info","info_pagamento","situacao"]:
        contrato.pop(k, None)

    cid = insert_contrato(contrato)

    # Parcelas
    header_row = rows[4]
    parc_col = next((ci for ci,cell in enumerate(header_row) if str(cell or "").strip().upper()=="PARCELA"), None)
    if parc_col:
        for row in rows[5:]:
            num = row[parc_col]
            if num is None: continue
            try: num_int = int(float(str(num)))
            except: continue
            if num_int <= 0: continue
            insert_parcela({
                "contrato_id": cid,
                "numero":       num_int,
                "data_emissao": _to_date(row[parc_col+1]),
                "valor":        _to_float(row[parc_col+2]),
                "saldo_atual":  _to_float(row[parc_col+3]),
                "situacao":     str(row[parc_col+4] or "").strip() or None,
                "numero_nf":    _safe_nf(row[parc_col+5]),
            })

    icon = "🟢" if status == "ATIVO" else "⚪"
    print(f"  {icon} {nome_aba:<25} {empresa_gg:<15} R${contrato['valor_total']:>12,.2f}")

wb.close()

# ══════════════════════════════════════════════════════════════════════════════
# PARTE 2 — Base estratégica de 30 contratos (fornecida manualmente)
# ══════════════════════════════════════════════════════════════════════════════
print("\n📋 Inserindo base estratégica de contratos...")

ESTRATEGICOS = [
    # (contratante, empresa_gg, tipo, categoria, servico, status, valor_total, valor_por_amostra,
    #  valor_recorrente, recorrente, renovacao_auto, internacional, white_label, tem_comissao,
    #  comissao_pct, lgpd, confidencialidade, nao_conc, prop_intl, internac_dados, observacoes)
    ("Novozymes",
     "GG SOLUÇÕES", "Prestação científica", "Metagenômica",
     "Biodiversidade + Illumina MiSeq + PCR",
     "ATIVO", 0, 440, 0, True, False, True, False, False, 0,
     True, True, False, False, True,
     "Contrato estratégico internacional. Volume por amostra R$390–490."),

    ("Omni",
     "GG SOLUÇÕES", "White label", "SaaS + genética humana",
     "Kits genéticos humanos/esportivos",
     "ATIVO", 0, 0, 0, True, False, False, True, False, 0,
     True, True, True, True, False,
     "Plataforma integrada white label."),

    ("SCI Agro",
     "GG SOLUÇÕES", "Implantação", "Consultoria técnica",
     "Implantação de laboratório molecular",
     "ENCERRADO", 10000, 0, 0, False, False, False, False, False, 0,
     False, False, False, False, False,
     "Consultoria + treinamento. Proposta 30 dias."),

    ("Total Bio / Biotrop",
     "GG SOLUÇÕES", "Prestação laboratorial", "Microbiologia",
     "Quantificação microbiológica — ensaios industriais",
     "ATIVO", 0, 0, 0, True, False, False, False, False, 0,
     False, True, False, False, False,
     "Ensaios industriais contínuos."),

    ("Towing",
     "GG SOLUÇÕES", "Prestação laboratorial", "Internacional",
     "Shotgun metagenomics",
     "ATIVO", 17800, 0, 17800, False, True, True, False, False, 0,
     True, True, False, True, True,
     "Cliente japonês. Contrato 12 meses com renovação automática."),

    ("Vittia",
     "GG SOLUÇÕES", "Cooperação técnica", "Pesquisa agrícola",
     "Bacillus subtilis — pesquisa financiada",
     "VENCIDO", 51000, 0, 0, False, False, False, False, False, 0,
     False, True, False, True, False,
     "Pesquisa financiada 12 meses. Possível vencido."),

    ("Biomcrop",
     "GG SOLUÇÕES", "Desenvolvimento científico", "Bioinsumos",
     "Microbiologia aplicada",
     "ATIVO", 0, 0, 0, False, False, False, False, False, 0,
     False, True, True, True, False,
     "Possui exclusividade contratual."),

    ("Biogenic",
     "GOSOLOS", "SaaS + laboratório", "Plataforma + análises",
     "GoSolos + microbioma",
     "ATIVO", 0, 0, 0, True, True, False, False, False, 0,
     True, True, False, True, False,
     "Plataforma SaaS ativa com análises recorrentes."),

    ("Agroceres / Binova",
     "GG SOLUÇÕES", "Desenvolvimento tecnológico", "qPCR",
     "Primers e sondas — transferência tecnológica",
     "ATIVO", 0, 0, 0, False, False, False, False, False, 0,
     False, True, False, True, False,
     "Desenvolvimento e transferência tecnológica."),

    ("CENCODERMA / Grupo Boticário",
     "GG SOLUÇÕES", "P&D", "Cooperação científica",
     "Desenvolvimento conjunto",
     "ATIVO", 0, 0, 0, False, False, False, False, False, 0,
     True, True, False, True, False,
     "PI compartilhada. Parceria estratégica Grupo Boticário."),

    ("Mosaic",
     "GOSOLOS", "Cooperação técnica", "Pesquisa microbiológica",
     "Estudos microbiológicos",
     "ATIVO", 0, 0, 0, False, False, True, False, False, 0,
     True, True, False, True, True,
     "Compliance internacional. Multinacional."),

    ("Elanco",
     "GG SOLUÇÕES", "Prestação corporativa", "Multinacional",
     "Serviços laboratoriais corporativos",
     "ATIVO", 0, 0, 0, False, False, True, False, False, 0,
     True, True, False, False, True,
     "Compliance FCPA. Multinacional."),

    ("ADAPAR",
     "GG SOLUÇÕES", "Governamental", "Diagnóstico",
     "Microbiologia — projeto público",
     "ENCERRADO", 0, 0, 0, False, False, False, False, False, 0,
     False, False, False, False, False,
     "Contrato governamental encerrado."),

    ("CIAT / TerraBio",
     "GOSOLOS", "Pesquisa internacional", "Metagenoma",
     "Metagenômica — cooperação internacional",
     "ATIVO", 0, 0, 0, False, False, True, False, False, 0,
     True, True, False, True, True,
     "Cooperação internacional."),

    ("Baldan",
     "GG SOLUÇÕES", "Prestação de serviço", "Laboratorial",
     "Análises microbiológicas",
     "ATIVO", 0, 0, 0, False, False, False, False, False, 0,
     False, True, False, False, False,
     "Contrato operacional simples."),

    ("Cosmocel",
     "GG SOLUÇÕES", "Cooperação científica", "Desenvolvimento",
     "Pesquisa microbiológica",
     "ATIVO", 0, 0, 0, False, False, True, False, False, 0,
     True, True, False, True, True,
     "PI e confidencialidade. Internacional."),

    ("JCO",
     "GG SOLUÇÕES", "Prestação recorrente", "Escala industrial",
     "Análises moleculares em escala",
     "ATIVO", 0, 0, 0, True, False, False, False, False, 0,
     False, True, False, False, False,
     "Alto volume recorrente."),

    ("Efense",
     "GOSOLOS", "SaaS + microbioma", "Plataforma GoSolos",
     "Microbioma + plataforma GoSolos",
     "ATIVO", 0, 1025, 0, True, True, False, False, False, 0,
     True, True, False, True, False,
     "Faixa R$750–1.300/amostra (escalonado). Compensação financeira."),

    ("Brandt",
     "GOSOLOS", "SaaS + microbioma", "Enterprise",
     "Microbioma do solo — enterprise",
     "ATIVO", 0, 0, 69600, True, True, False, False, False, 0,
     True, True, False, True, False,
     "Até R$69.600/ano. Volume + bônus. Possui aditivo."),

    ("BlackBio",
     "GG SOLUÇÕES", "SaaS + laboratório", "SaaS + microbioma",
     "Microbioma + plataforma SaaS",
     "ATIVO", 19220, 0, 0, False, False, False, False, False, 0,
     True, True, False, True, False,
     "R$19.220 pacote anual."),

    ("Danco",
     "GG SOLUÇÕES", "Prestação pontual", "Projeto fechado",
     "Análises microbiológicas",
     "ENCERRADO", 8300, 0, 0, False, False, False, False, False, 0,
     False, True, False, False, False,
     "Projeto 180 dias. Sem recorrência."),

    ("CTC",
     "GG SOLUÇÕES", "Compliance corporativo", "Corporativo",
     "Compliance + prestação laboratorial",
     "ATIVO", 0, 0, 0, True, False, False, False, False, 0,
     True, True, True, False, False,
     "Forte compliance corporativo."),

    ("Indigo",
     "GOSOLOS", "SaaS + microbioma", "Enterprise",
     "Microbioma do solo — enterprise",
     "ATIVO", 25500, 0, 25500, True, False, False, False, False, 0,
     True, True, False, True, False,
     "Excedentes e bônus incluídos."),

    ("Semiocrop",
     "GOSOLOS", "Prestação contínua", "Pay per use",
     "Microbioma do solo — consumo livre",
     "ATIVO", 0, 710, 0, True, True, False, False, False, 0,
     True, True, False, False, False,
     "R$710/amostra. Sem mínimo. Renovável 6 meses."),

    ("Sustentar Agro",
     "GOSOLOS", "SaaS + laboratório", "Enterprise",
     "Microbioma + plataforma SaaS",
     "ATIVO", 65000, 0, 0, False, False, False, True, False, 0,
     True, True, False, True, False,
     "100 amostras incluídas. Possível white label."),

    ("TecSolo",
     "GOSOLOS", "Representação comercial", "Canal/parceria",
     "Representação microbioma do solo",
     "ATIVO", 0, 0, 0, False, True, False, False, True, 12.0,
     False, True, True, False, False,
     "Comissão 10–14% escalonada. Canal comercial."),

    ("Tradecorp",
     "GG SOLUÇÕES", "Comodato", "Equipamentos",
     "Comodato operacional de equipamentos",
     "ATIVO", 0, 0, 0, False, True, False, False, False, 0,
     False, True, False, False, False,
     "Controle patrimonial. Já renovado."),
]

for row in ESTRATEGICOS:
    (contratante, empresa_gg, tipo, categoria, servico, status,
     valor_total, valor_por_amostra, valor_recorrente, recorrente, renovacao_auto,
     internacional, white_label, tem_comissao, comissao_pct,
     lgpd, confidencialidade, nao_conc, prop_intl, internac_dados, obs) = row

    insert_contrato({
        "contratante":            contratante,
        "empresa_gg":             empresa_gg,
        "tipo_contrato":          tipo,
        "categoria":              categoria,
        "servico_principal":      servico,
        "status":                 status,
        "valor_total":            valor_total,
        "valor_por_amostra":      valor_por_amostra,
        "valor_recorrente":       valor_recorrente,
        "recorrente":             1 if recorrente else 0,
        "renovacao_automatica":   1 if renovacao_auto else 0,
        "internacional":          1 if internacional else 0,
        "white_label":            1 if white_label else 0,
        "tem_comissao":           1 if tem_comissao else 0,
        "comissao_percentual":    comissao_pct,
        "lgpd":                   1 if lgpd else 0,
        "confidencialidade":      1 if confidencialidade else 0,
        "nao_concorrencia":       1 if nao_conc else 0,
        "propriedade_intelectual":1 if prop_intl else 0,
        "internacionalizacao_dados": 1 if internac_dados else 0,
        "observacoes":            obs,
    })
    icon = "🟢" if status == "ATIVO" else ("⚪" if status == "ENCERRADO" else "🟡")
    print(f"  {icon} {contratante:<30} {empresa_gg:<15} {tipo}")

# ── Resumo final ──────────────────────────────────────────────────────────────
from db_contratos_sqlite import resumo_contratos
res = resumo_contratos()
print(f"""
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
🎉 Seed concluído!
   Total de contratos  : {res['qtd_ativos'] + res['qtd_encerrados'] + res['qtd_vencidos']}
   Ativos              : {res['qtd_ativos']}
   Encerrados/vencidos : {res['qtd_encerrados'] + res['qtd_vencidos']}
   Internacionais      : {res['qtd_internacionais']}
   White label         : {res['qtd_white_label']}
   Recorrentes         : {res['qtd_recorrentes']}
   Renovação auto      : {res['qtd_renovacao_auto']}
   Receita recorrente  : R$ {res['receita_recorrente']:,.2f}/ano
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
""")
