"""Página 8 — Fluxo de Caixa · Todas as Empresas."""
import io
import json
from datetime import date, timedelta
from pathlib import Path

import openpyxl
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

from utils import (GLOBAL_CSS, BRAND, brl, soma, kpi_card, plotly_layout,
                   sidebar_header, get_empresas_disponiveis,
                   load_data_unificado, load_vencidas_unificado,
                   load_companies_data, load_companies_vencidas,
                   tabela_marcavel)

CONTRATOS_XLS  = Path(__file__).parent.parent / "data" / "contratos.xlsx"
CONTRATOS_JSON = Path(__file__).parent.parent / "data" / "contratos_manual.json"


@st.cache_data(ttl=300, show_spinner=False)
def load_parcelas_espera(caminho: str) -> list:
    """Retorna parcelas EM ESPERA de todos os contratos como itens de fluxo."""
    resultado = []
    if not Path(caminho).exists():
        return resultado
    try:
        wb = openpyxl.load_workbook(caminho, data_only=True)
        for nome_aba in wb.sheetnames:
            if nome_aba == "RESUMO":
                continue
            ws    = wb[nome_aba]
            rows  = list(ws.iter_rows(values_only=True))

            # Extrai contratante e empresa das primeiras linhas
            contratante, empresa = nome_aba, "—"
            for r in rows:
                chave = str(r[1]).strip().upper() if len(r) > 1 and r[1] else ""
                valor = r[2] if len(r) > 2 else None
                if chave == "CONTRATANTE" and valor:
                    contratante = str(valor).strip()
                if chave == "EMPRESA" and valor:
                    empresa = str(valor).strip()

            # Localiza cabeçalho de parcelas
            ph_row = None
            for i, r in enumerate(rows):
                if (any(str(v).strip().upper() == "PARCELA" for v in r if v) and
                        any(str(v).strip().upper()[:5] in ("EMISS", "VALOR") for v in r if v)):
                    ph_row = i
                    break
            if ph_row is None:
                continue

            header = [str(v).strip().upper() if v else "" for v in rows[ph_row]]
            def ci(kws):
                for kw in kws:
                    for i, h in enumerate(header):
                        if kw in h: return i
                return None

            cp  = ci(["PARCELA"])
            cem = ci(["EMISS"])
            cvl = ci(["VALOR"])
            cst = ci(["SITUA"])

            for r in rows[ph_row + 1:]:
                num = r[cp] if cp is not None and cp < len(r) else None
                if not isinstance(num, (int, float)):
                    continue
                sit = str(r[cst]).strip().upper() if cst is not None and cst < len(r) and r[cst] else "EM ESPERA"
                if sit != "EM ESPERA":
                    continue
                emissao = r[cem] if cem is not None and cem < len(r) else None
                valor   = r[cvl] if cvl is not None and cvl < len(r) else None
                if not isinstance(valor, (int, float)) or not emissao:
                    continue

                # Converte data
                if isinstance(emissao, (date, datetime)):
                    dt_str = emissao.strftime("%Y-%m-%d") if hasattr(emissao, 'strftime') else str(emissao)[:10]
                else:
                    dt_str = str(emissao)[:10]

                resultado.append({
                    "valor":       float(valor),
                    "dtVenc":      dt_str,
                    "nomeContato": contratante,
                    "empresa":     empresa,
                    "descricao":   f"Contrato {nome_aba} — Parcela {int(num)}",
                    "_origem":     "contrato",
                })
    except Exception as e:
        pass

    # Parcelas de contratos manuais
    if CONTRATOS_JSON.exists():
        try:
            manual = json.loads(CONTRATOS_JSON.read_text(encoding="utf-8"))
            for aba, cont in manual.get("novos_contratos", {}).items():
                contratante = cont.get("info", {}).get("CONTRATANTE", aba)
                empresa     = cont.get("info", {}).get("EMPRESA", "—")
                for p in cont.get("parcelas", []):
                    if str(p.get("Situação", "")).upper() != "EM ESPERA":
                        continue
                    emissao = p.get("Emissão", "")
                    valor   = p.get("Valor")
                    if not valor or not emissao:
                        continue
                    resultado.append({
                        "valor":       float(valor),
                        "dtVenc":      str(emissao)[:10],
                        "nomeContato": contratante,
                        "empresa":     empresa,
                        "descricao":   f"Contrato {aba} — Parcela {p.get('Parcela','')}",
                        "_origem":     "contrato",
                    })
        except Exception:
            pass

    return resultado

st.set_page_config(page_title="Fluxo de Caixa | GoGenetic", page_icon="💰", layout="wide")
st.markdown(GLOBAL_CSS, unsafe_allow_html=True)

# ── Sidebar ────────────────────────────────────────────────────────────────────
sidebar_header()

with st.sidebar:
    _disponiveis   = get_empresas_disponiveis()
    empresa_opcoes = ["Todas"] + _disponiveis
    empresa_sel    = st.selectbox("🏢 Empresa", empresa_opcoes)
    empresas_ativas = _disponiveis if empresa_sel == "Todas" else [empresa_sel]

    st.markdown("---")
    st.markdown("**📅 Horizonte**")
    horizonte = st.selectbox("Período futuro", [
        "Próximos 30 dias",
        "Próximos 60 dias",
        "Próximos 90 dias",
        "Próximos 6 meses",
        "Este ano",
    ], index=1)

    MAP_HORIZONTE = {
        "Próximos 30 dias":  30,
        "Próximos 60 dias":  60,
        "Próximos 90 dias":  90,
        "Próximos 6 meses":  180,
        "Este ano":          (date(date.today().year, 12, 31) - date.today()).days,
    }
    dias_horizonte = MAP_HORIZONTE[horizonte]
    dt_ini = date.today()
    dt_fim = date.today() + timedelta(days=dias_horizonte)

    incluir_pagar = st.checkbox("Incluir A Pagar (saldo líquido)", value=True)

    st.markdown("---")
    if st.button("🔄 Atualizar dados", use_container_width=True):
        st.cache_data.clear()
        st.rerun()
    st.caption("⏱ Cache: 5 min")

# ── Carrega dados (empresas em paralelo) ───────────────────────────────────────
with st.spinner("Carregando fluxo de caixa..."):
    _dt_ini_s = dt_ini.strftime("%Y-%m-%d")
    _dt_fim_s = dt_fim.strftime("%Y-%m-%d")
    _dados_map = load_companies_data(empresas_ativas, _dt_ini_s, _dt_fim_s)
    _venc_map  = load_companies_vencidas(empresas_ativas)

    rec_items, pag_items = [], []
    venc_rec, venc_pag   = [], []
    for nome in empresas_ativas:
        dados = _dados_map[nome]
        for it in dados["contas_receber"]: rec_items.append({**it, "empresa": nome})
        for it in dados["contas_pagar"]:   pag_items.append({**it, "empresa": nome})
        v = _venc_map[nome]
        for it in v["receber"]: venc_rec.append({**it, "empresa": nome})
        for it in v["pagar"]:   venc_pag.append({**it, "empresa": nome})

    # Parcelas de contratos EM ESPERA (não faturadas, fora do eGestor)
    todas_parcelas = load_parcelas_espera(str(CONTRATOS_XLS))
    dt_ini_str = dt_ini.strftime("%Y-%m-%d")
    dt_fim_str = dt_fim.strftime("%Y-%m-%d")
    parcelas_periodo = [
        p for p in todas_parcelas
        if dt_ini_str <= str(p.get("dtVenc", ""))[:10] <= dt_fim_str
    ]
    rec_items += parcelas_periodo
    total_contratos = soma(parcelas_periodo, "valor")

total_rec  = soma(rec_items,  "valor")
total_pag  = soma(pag_items,  "valor")
total_vr   = soma(venc_rec,   "valor")
total_vp   = soma(venc_pag,   "valor")
saldo_liq  = total_rec - (total_pag if incluir_pagar else 0)

# ── Header ─────────────────────────────────────────────────────────────────────
empresa_label = empresas_ativas[0] if len(empresas_ativas) == 1 else "Grupo GoGenetic"
st.markdown(f"""
<p class="page-title">💰 Fluxo de Caixa</p>
<p class="page-sub">{empresa_label} · até {dt_fim.strftime('%d/%m/%Y')}</p>
""", unsafe_allow_html=True)

# ── KPIs ───────────────────────────────────────────────────────────────────────
c1, c2, c3, c4, c5, c6 = st.columns(6)
kpi_card(c1, "📥", "A Receber",      brl(total_rec - total_contratos), f"{len(rec_items) - len(parcelas_periodo)} títulos", border="#10B981")
kpi_card(c2, "📋", "Contratos",      brl(total_contratos), f"{len(parcelas_periodo)} parcelas em espera", border="#7E16B8")
kpi_card(c3, "⚠️", "Vencido a Rec.", brl(total_vr),  f"{len(venc_rec)} títulos",  border="#F59E0B",
         value_class="kpi-warn" if total_vr > 0 else "")
kpi_card(c4, "📤", "A Pagar",        brl(total_pag), f"{len(pag_items)} títulos", border="#EF4444")
kpi_card(c5, "🚨", "Vencido a Pag.", brl(total_vp),  f"{len(venc_pag)} títulos",  border="#EF4444",
         value_class="kpi-negative" if total_vp > 0 else "")
kpi_card(c6, "💼", "Saldo Líquido",  brl(saldo_liq), f"receber+contratos − {'pagar' if incluir_pagar else 'zero'}",
         border="#7E16B8" if saldo_liq >= 0 else "#EF4444",
         value_class="kpi-positive" if saldo_liq >= 0 else "kpi-negative")

# ── Prepara DataFrame ──────────────────────────────────────────────────────────
def prep_items(items: list, tipo: str) -> pd.DataFrame:
    if not items:
        return pd.DataFrame()
    df = pd.DataFrame(items)
    df["valor"]  = pd.to_numeric(df["valor"], errors="coerce").fillna(0)
    df["dtVenc"] = pd.to_datetime(df["dtVenc"], errors="coerce")
    df = df.dropna(subset=["dtVenc"])
    df["tipo"]   = tipo
    df["semana"] = df["dtVenc"].dt.to_period("W").apply(lambda p: p.start_time.date())
    df["mes"]    = df["dtVenc"].dt.to_period("M").astype(str)
    return df

df_rec  = prep_items(rec_items,  "A Receber")
df_pag  = prep_items(pag_items,  "A Pagar")
df_vr   = prep_items(venc_rec,   "Vencido Rec.")
df_vp   = prep_items(venc_pag,   "Vencido Pag.")

# ── Gráfico Semanal ────────────────────────────────────────────────────────────
st.markdown("<div class='section-title'>Fluxo Semanal</div>", unsafe_allow_html=True)

frames = []
if not df_rec.empty:  frames.append(df_rec.assign(sinal= 1))
if not df_vr.empty:   frames.append(df_vr.assign(sinal= 1))
if incluir_pagar:
    if not df_pag.empty:  frames.append(df_pag.assign(sinal=-1))
    if not df_vp.empty:   frames.append(df_vp.assign(sinal=-1))

if frames:
    df_all = pd.concat(frames, ignore_index=True)
    df_all["valor_sig"] = df_all["valor"] * df_all["sinal"]

    # Agrupado por semana e tipo
    df_sem = (df_all.groupby(["semana", "tipo"])["valor"]
              .sum().reset_index()
              .sort_values("semana"))

    cor_map = {
        "A Receber":    "#10B981",
        "Vencido Rec.": "#F59E0B",
        "A Pagar":      "#EF4444",
        "Vencido Pag.": "#7F1D1D",
    }

    fig = px.bar(
        df_sem, x="semana", y="valor", color="tipo",
        barmode="group",
        color_discrete_map=cor_map,
        labels={"semana": "", "valor": "R$", "tipo": ""},
    )

    # Linha de saldo acumulado líquido
    df_liq = (df_all.groupby("semana")["valor_sig"]
              .sum().reset_index()
              .sort_values("semana"))
    df_liq["acumulado"] = df_liq["valor_sig"].cumsum()

    fig.add_trace(go.Scatter(
        x=df_liq["semana"],
        y=df_liq["acumulado"],
        name="Saldo Acumulado",
        mode="lines+markers",
        line=dict(color="#7E16B8", width=2.5, dash="dot"),
        marker=dict(size=6),
        yaxis="y2",
    ))
    fig.update_layout(
        yaxis2=dict(
            overlaying="y", side="right",
            showgrid=False,
            tickfont=dict(color="#7E16B8", size=10),
            title=dict(text="Saldo Acum.", font=dict(color="#7E16B8", size=10)),
        )
    )
    plotly_layout(fig)
    fig.update_layout(height=380)
    st.plotly_chart(fig, use_container_width=True)

# ── Gráfico por Empresa ────────────────────────────────────────────────────────
if len(empresas_ativas) > 1 and not df_rec.empty:
    st.markdown("<div class='section-title'>A Receber por Empresa</div>", unsafe_allow_html=True)
    df_emp = df_rec.groupby("empresa")["valor"].sum().reset_index().sort_values("valor", ascending=False)
    from utils import BRAND
    df_emp["cor"] = df_emp["empresa"].map(lambda e: BRAND.get(e, {}).get("primary", "#7E16B8"))
    fig_emp = px.bar(
        df_emp, x="empresa", y="valor",
        color="empresa",
        color_discrete_map={e: BRAND.get(e, {}).get("primary", "#7E16B8") for e in df_emp["empresa"]},
        labels={"empresa": "", "valor": "R$"},
        text=df_emp["valor"].apply(brl),
    )
    fig_emp.update_traces(textposition="outside")
    plotly_layout(fig_emp)
    st.plotly_chart(fig_emp, use_container_width=True)

# ── Tabelas ────────────────────────────────────────────────────────────────────
n_egst  = len(rec_items) - len(parcelas_periodo)
n_contr = len(parcelas_periodo)
tab_rec_t, tab_contr_t, tab_pag_t, tab_venc_t = st.tabs([
    f"📥 A Receber eGestor ({n_egst})",
    f"📋 Contratos em Espera ({n_contr})",
    f"📤 A Pagar ({len(pag_items)})",
    f"⚠️ Vencidos ({len(venc_rec) + len(venc_pag)})",
])

COLS_REC = {"empresa": "Empresa", "dtVenc": "Vencimento", "nomeContato": "Cliente",
            "valor": "Valor", "descricao": "Descrição"}
COLS_PAG = {"empresa": "Empresa", "dtVenc": "Vencimento", "nomeContato": "Fornecedor",
            "valor": "Valor", "descricao": "Descrição"}


def tabela_fluxo(items: list, cols_map: dict, key: str):
    if not items:
        st.info("Nenhum registro.")
        return
    df = pd.DataFrame(items)
    df["valor"]  = pd.to_numeric(df["valor"], errors="coerce").fillna(0)
    df["dtVenc"] = pd.to_datetime(df["dtVenc"], errors="coerce")
    df = df.sort_values("dtVenc")

    existing = [c for c in cols_map if c in df.columns]
    df_show  = df[existing].rename(columns=cols_map).copy()
    if "Vencimento" in df_show.columns:
        df_show["Vencimento"] = pd.to_datetime(df_show["Vencimento"], errors="coerce").dt.strftime("%d/%m/%Y")

    df_num = df.copy()

    # Coluna de dias até vencimento
    if "dtVenc" in df.columns:
        hoje = pd.Timestamp.today().normalize()
        df_show["Dias"] = (df["dtVenc"] - hoje).dt.days.fillna(0).astype(int)

    df_show["Valor"] = df_num["valor"].apply(brl)

    sel, _ = tabela_marcavel(df_show, key=key)

    col_i, col_s, col_e = st.columns([2, 2, 1])
    with col_i:
        st.caption(f"{'✅ ' + str(len(sel)) + ' selecionado(s) de ' if sel else 'Total: '}{len(items)} registros")
    with col_s:
        val   = df_num.iloc[sel]["valor"].sum() if sel else df_num["valor"].sum()
        label = "Soma selecionados" if sel else "Total"
        borda = "#7E16B8" if sel else "rgba(126,22,184,0.2)"
        st.markdown(
            f"<div style='background:#F5F0FA;border:1px solid {borda};border-radius:8px;"
            f"padding:8px 16px;text-align:right'>"
            f"<span style='font-size:.72rem;color:#8B6BAE;text-transform:uppercase;letter-spacing:1px'>{label}</span><br>"
            f"<span style='font-size:1.3rem;font-weight:700;color:#1A1033'>{brl(val)}</span>"
            f"</div>", unsafe_allow_html=True,
        )
    with col_e:
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as w:
            (df_show.iloc[sel] if sel else df_show).to_excel(w, index=False, sheet_name="Fluxo")
        st.download_button(
            f"📥 Excel" + (f" ({len(sel)} sel.)" if sel else ""),
            buf.getvalue(),
            file_name=f"fluxo_{key}_{date.today():%Y%m%d}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True, key=f"dl_{key}",
        )


with tab_rec_t:
    egst_items = [it for it in rec_items if it.get("_origem") != "contrato"]
    tabela_fluxo(egst_items, COLS_REC, "receber")

with tab_contr_t:
    if parcelas_periodo:
        COLS_CONTR = {"empresa": "Empresa", "dtVenc": "Vencimento",
                      "nomeContato": "Contratante", "valor": "Valor", "descricao": "Descrição"}
        tabela_fluxo(parcelas_periodo, COLS_CONTR, "contratos")
    else:
        st.info("Nenhuma parcela de contrato em espera para o período selecionado.")

with tab_pag_t:
    tabela_fluxo(pag_items, COLS_PAG, "pagar")

with tab_venc_t:
    st.markdown("#### 📥 Vencidos a Receber")
    tabela_fluxo(venc_rec, COLS_REC, "venc_rec")
    st.markdown("#### 📤 Vencidos a Pagar")
    tabela_fluxo(venc_pag, COLS_PAG, "venc_pag")
