"""Página 5 — Contratos."""
import json
import io
from pathlib import Path
from datetime import date, datetime, timedelta

import openpyxl
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

from utils import GLOBAL_CSS, brl, kpi_card, plotly_layout, sidebar_header

st.set_page_config(page_title="Contratos | GoGenetic", page_icon="📋", layout="wide")
st.markdown(GLOBAL_CSS, unsafe_allow_html=True)

DATA_FILE    = Path(__file__).parent.parent / "data" / "contratos.xlsx"
MANUAL_FILE  = Path(__file__).parent.parent / "data" / "contratos_manual.json"

SITUACOES = ["EM ESPERA", "FATURADO", "ENTREGUE", "FALTA", "CANCELADO"]
EMPRESAS  = ["GG SOLUÇÕES", "GOSOLOS", "GOGENETIC PESQUISA", "GOGENETIC YOU"]

# ── Helpers de persistência ────────────────────────────────────────────────────

def load_manual() -> dict:
    if MANUAL_FILE.exists():
        return json.loads(MANUAL_FILE.read_text(encoding="utf-8"))
    return {"overrides": {}, "novos_contratos": {}}


def save_manual(data: dict):
    MANUAL_FILE.parent.mkdir(exist_ok=True)
    MANUAL_FILE.write_text(json.dumps(data, indent=2, ensure_ascii=False, default=str),
                            encoding="utf-8")


# ── Sidebar ────────────────────────────────────────────────────────────────────
sidebar_header()

with st.sidebar:
    st.markdown("**📁 Arquivo de Contratos**")
    uploaded = st.file_uploader("Substituir planilha", type=["xlsx"])
    if uploaded:
        DATA_FILE.parent.mkdir(exist_ok=True)
        DATA_FILE.write_bytes(uploaded.read())
        st.success("✅ Arquivo atualizado!")
        st.cache_data.clear()
        st.rerun()

    if DATA_FILE.exists():
        st.caption(f"📄 {DATA_FILE.name}")
        st.caption(f"Atualizado: {date.fromtimestamp(DATA_FILE.stat().st_mtime).strftime('%d/%m/%Y')}")

    st.markdown("---")
    if st.button("🔄 Recarregar dados", use_container_width=True):
        st.cache_data.clear()
        st.rerun()

# ── Carrega Excel ──────────────────────────────────────────────────────────────
if not DATA_FILE.exists():
    st.warning("⚠️ Nenhum arquivo encontrado. Use o menu lateral para enviar a planilha.")
    st.stop()


@st.cache_data(ttl=300, show_spinner=False)
def load_excel(caminho: str):
    wb = openpyxl.load_workbook(caminho, data_only=True)

    # ── RESUMO ────────────────────────────────────────────────────────────────
    ws   = wb["RESUMO"]
    rows = list(ws.iter_rows(values_only=True))

    ativos, encerrados = [], []
    secao = None
    for idx_r, row in enumerate(rows):
        vals = list(row)
        if any(str(v).strip().upper() == "CONTRATANTE" for v in vals if v):
            prev = [str(v) for v in rows[idx_r - 1] if v] if idx_r > 0 else []
            secao = "enc" if any("ENCERRADO" in p.upper() for p in prev) else "atv"
            continue
        if any(str(v).strip().upper() == "ENCERRADOS" for v in vals if v):
            secao = "enc"
            continue

        empresa    = str(vals[2]).strip() if len(vals) > 2 and vals[2] else ""
        contratante= str(vals[3]).strip() if len(vals) > 3 and vals[3] else ""
        if not empresa or not contratante or empresa.upper() in ("EMPRESA", ""):
            continue

        def to_pct(v):
            if v is None: return None
            if isinstance(v, (int, float)): return round(float(v) * 100, 1)
            return str(v)

        linha = {
            "Empresa":     empresa,
            "Contratante": contratante,
            "Parcela":     int(vals[4]) if isinstance(vals[4], (int, float)) else vals[4],
            "Data Pgto":   str(vals[5]).strip() if vals[5] else "—",
            "Entregue":    to_pct(vals[6]),
            "Faturado":    to_pct(vals[7]),
            "Contrato":    float(vals[8]) if isinstance(vals[8], (int, float)) else 0.0,
            "Saldo":       float(vals[9]) if isinstance(vals[9], (int, float)) else 0.0,
            "Situação":    "Encerrado" if secao == "enc" else "Ativo",
        }
        (encerrados if secao == "enc" else ativos).append(linha)

    # ── Abas individuais ──────────────────────────────────────────────────────
    detalhes = {}
    for nome_aba in wb.sheetnames:
        if nome_aba == "RESUMO":
            continue
        try:
            ws2   = wb[nome_aba]
            rows2 = list(ws2.iter_rows(values_only=True))

            info = {}
            for r in rows2:
                chave = str(r[1]).strip().upper() if len(r) > 1 and r[1] else ""
                valor = r[2] if len(r) > 2 else None
                if chave in ("EMPRESA","CONTRATANTE","RENOVAÇÃO","ASSINATURA","VALIDADE",
                             "KIT","AMOSTRAS","CONTRATO","PARCELA","SITUAÇÃO",
                             "INFORMAÇÕES","CONSUMO","SALDO","OBS"):
                    info[chave] = valor

            parcelas = []
            ph_row   = None
            for i, r in enumerate(rows2):
                if (any(str(v).strip().upper() == "PARCELA" for v in r if v) and
                        any(str(v).strip().upper()[:5] in ("EMISS", "VALOR") for v in r if v)):
                    ph_row = i
                    break

            if ph_row is not None:
                header = [str(v).strip().upper() if v else "" for v in rows2[ph_row]]
                def ci(kws):
                    for kw in kws:
                        for i, h in enumerate(header):
                            if kw in h: return i
                    return None

                cp  = ci(["PARCELA"])
                cem = ci(["EMISS"])
                cvl = ci(["VALOR"])
                csd = ci(["SALDO"])
                cst = ci(["SITUA"])
                cnf = ci(["NF"])

                for r in rows2[ph_row + 1:]:
                    num = r[cp] if cp is not None and cp < len(r) else None
                    if not isinstance(num, (int, float)): continue
                    parcelas.append({
                        "Parcela":  int(num),
                        "Emissão":  r[cem] if cem is not None and cem < len(r) else None,
                        "Valor":    float(r[cvl]) if cvl is not None and cvl < len(r) and isinstance(r[cvl], (int,float)) else None,
                        "Saldo":    float(r[csd]) if csd is not None and csd < len(r) and isinstance(r[csd], (int,float)) else None,
                        "Situação": str(r[cst]).strip() if cst is not None and cst < len(r) and r[cst] else "EM ESPERA",
                        "NF":       str(r[cnf]).strip() if cnf is not None and cnf < len(r) and r[cnf] else "",
                    })

            detalhes[nome_aba] = {"info": info, "parcelas": parcelas}
        except Exception:
            pass

    return ativos, encerrados, detalhes


with st.spinner("Carregando contratos..."):
    ativos_excel, encerrados_excel, detalhes_excel = load_excel(str(DATA_FILE))

# Merge com dados manuais
manual        = load_manual()
novos         = manual.get("novos_contratos", {})
overrides     = manual.get("overrides", {})

# Junta todos os detalhes (Excel + manuais)
detalhes_todos = {**detalhes_excel, **novos}

# Ativos = Excel ativos + contratos manuais com status Ativo
ativos_manual = [
    {
        "Empresa":     v["info"].get("EMPRESA", "—"),
        "Contratante": v["info"].get("CONTRATANTE", "—"),
        "Parcela":     len(v["parcelas"]),
        "Data Pgto":   v["info"].get("DATA_PGTO", "—"),
        "Entregue":    None,
        "Faturado":    None,
        "Contrato":    float(v["info"].get("CONTRATO", 0) or 0),
        "Saldo":       float(v["info"].get("SALDO", 0) or 0),
        "Situação":    "Ativo",
        "_manual":     True,
        "_aba":        k,
    }
    for k, v in novos.items()
]

todos_ativos    = ativos_excel + ativos_manual
todos_encerrados= encerrados_excel

df_ativos = pd.DataFrame(todos_ativos)
df_enc    = pd.DataFrame(todos_encerrados)

# ── Header ─────────────────────────────────────────────────────────────────────
st.markdown("""
<p class="page-title">📋 Contratos</p>
<p class="page-sub">Controle de contratos ativos, parcelas e baixas</p>
""", unsafe_allow_html=True)

# ── KPIs ───────────────────────────────────────────────────────────────────────
valor_total     = df_ativos["Contrato"].sum() if not df_ativos.empty else 0
saldo_total     = df_ativos["Saldo"].sum()    if not df_ativos.empty else 0
consumido_total = valor_total - saldo_total
pct_consumido   = (consumido_total / valor_total * 100) if valor_total else 0

c1, c2, c3, c4 = st.columns(4)
kpi_card(c1, "📋", "Contratos Ativos", str(len(df_ativos)),    f"{len(df_enc)} encerrados")
kpi_card(c2, "💰", "Valor Total",      brl(valor_total),       "soma contratos ativos")
kpi_card(c3, "✅", "Consumido",        brl(consumido_total),   f"{pct_consumido:.1f}% do total",
         value_class="kpi-positive")
kpi_card(c4, "📦", "Saldo Restante",   brl(saldo_total),       "a consumir/receber",
         border="rgba(245,166,35,0.4)", value_class="kpi-warn" if saldo_total > 0 else "")

if valor_total > 0:
    pct = min(pct_consumido, 100)
    cor = "#24B78C" if pct >= 80 else ("#F5A623" if pct >= 40 else "#7E16B8")
    st.markdown(f"""
    <div style='background:#F5F0FA;border-radius:10px;padding:12px 20px;margin:12px 0'>
      <div style='display:flex;justify-content:space-between;margin-bottom:6px'>
        <span style='font-size:.72rem;font-weight:600;color:#7E16B8;
          text-transform:uppercase;letter-spacing:1px'>Consumo Geral</span>
        <span style='font-size:.82rem;font-weight:700;color:{cor}'>{pct_consumido:.1f}%</span>
      </div>
      <div style='background:#E0D4F0;border-radius:6px;height:10px'>
        <div style='background:{cor};width:{pct:.1f}%;height:100%;border-radius:6px'></div>
      </div>
    </div>""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
tab_ativos, tab_baixas, tab_novo, tab_enc_tab = st.tabs([
    f"✅ Ativos ({len(df_ativos)})",
    "📝 Parcelas & Baixas",
    "➕ Novo Contrato",
    f"🗂️ Encerrados ({len(df_enc)})",
])

# ── Tab Ativos ─────────────────────────────────────────────────────────────────
with tab_ativos:
    if df_ativos.empty:
        st.info("Nenhum contrato ativo.")
    else:
        st.markdown("<div class='section-title'>Contratos em Vigência</div>", unsafe_allow_html=True)
        fig_sal = px.bar(
            df_ativos.sort_values("Saldo", ascending=False),
            x="Contratante", y=["Contrato", "Saldo"],
            barmode="overlay",
            color_discrete_map={"Contrato": "rgba(126,22,184,0.2)", "Saldo": "#7E16B8"},
            labels={"value": "R$", "variable": ""},
        )
        plotly_layout(fig_sal, "Valor Total vs Saldo")
        st.plotly_chart(fig_sal, use_container_width=True)

        df_show = df_ativos.copy()
        df_show["Contrato"] = df_show["Contrato"].apply(brl)
        df_show["Saldo"]    = df_show["Saldo"].apply(brl)
        df_show["Entregue"] = df_show["Entregue"].apply(
            lambda v: f"{v}%" if isinstance(v, (int, float)) else (str(v) if v else "—"))
        df_show["Faturado"] = df_show["Faturado"].apply(
            lambda v: f"{v}%" if isinstance(v, (int, float)) else (str(v) if v else "—"))
        cols_show = [c for c in ["Empresa","Contratante","Parcela","Data Pgto",
                                  "Entregue","Faturado","Contrato","Saldo"] if c in df_show.columns]
        st.dataframe(df_show[cols_show], use_container_width=True, hide_index=True)

# ── Tab Parcelas & Baixas ──────────────────────────────────────────────────────
with tab_baixas:
    abas_lista = list(detalhes_todos.keys())
    if not abas_lista:
        st.info("Nenhum contrato encontrado.")
    else:
        contrato_sel = st.selectbox("📄 Selecione o contrato", abas_lista, key="sel_baixas")
        det  = detalhes_todos[contrato_sel]
        info = det["info"]

        # Dados do contrato (colapsável)
        with st.expander("📌 Dados do contrato", expanded=False):
            col_a, col_b, col_c = st.columns(3)
            campos_info = [
                ("Empresa",     info.get("EMPRESA",     info.get("empresa", "—"))),
                ("Contratante", info.get("CONTRATANTE", info.get("contratante", "—"))),
                ("Validade",    info.get("VALIDADE",    info.get("validade", "—"))),
                ("Kit",         info.get("KIT",         info.get("kit", "—"))),
                ("Amostras",    info.get("AMOSTRAS",    info.get("amostras", "—"))),
                ("Valor Total", brl(info["CONTRATO"]) if isinstance(info.get("CONTRATO"), (int,float)) else str(info.get("CONTRATO","—"))),
                ("Parcela",     brl(info["PARCELA"])   if isinstance(info.get("PARCELA"),  (int,float)) else str(info.get("PARCELA","—"))),
                ("Situação",    info.get("SITUAÇÃO",    info.get("situacao", "—"))),
                ("Obs",         info.get("OBS",         info.get("obs", "—"))),
            ]
            for i, (lbl, val) in enumerate(campos_info):
                col = [col_a, col_b, col_c][i % 3]
                data_str = val.strftime("%d/%m/%Y") if hasattr(val, "strftime") else str(val)
                col.markdown(f"**{lbl}:** {data_str}")

        # Parcelas com editor
        parcelas_base = det.get("parcelas", [])
        ov = overrides.get(contrato_sel, {})

        # Aplica overrides existentes
        parcelas_merged = []
        for p in parcelas_base:
            num = str(p["Parcela"])
            if num in ov:
                parcelas_merged.append({**p, **ov[num]})
            else:
                parcelas_merged.append(p.copy())

        if not parcelas_merged:
            st.info("Nenhuma parcela cadastrada neste contrato.")
        else:
            df_edit = pd.DataFrame(parcelas_merged)

            # Formata datas para exibição
            df_edit["Emissão"] = pd.to_datetime(df_edit["Emissão"], errors="coerce").dt.date

            st.markdown("<div class='section-title'>Parcelas — edite situação, data e NF diretamente na tabela</div>",
                        unsafe_allow_html=True)
            st.caption("✏️ Clique em qualquer célula para editar · Situação, Emissão e NF são editáveis")

            edited = st.data_editor(
                df_edit,
                use_container_width=True,
                hide_index=True,
                num_rows="fixed",
                column_config={
                    "Parcela":  st.column_config.NumberColumn("Parc.", width="small", disabled=True),
                    "Emissão":  st.column_config.DateColumn("Emissão", format="DD/MM/YYYY"),
                    "Valor":    st.column_config.NumberColumn("Valor R$", format="R$ %.2f", disabled=True),
                    "Saldo":    st.column_config.NumberColumn("Saldo R$", format="R$ %.2f", disabled=True),
                    "Situação": st.column_config.SelectboxColumn("Situação", options=SITUACOES),
                    "NF":       st.column_config.TextColumn("NF"),
                },
            )

            # Botão salvar baixas
            col_sv, col_info2 = st.columns([1, 3])
            with col_sv:
                if st.button("💾 Salvar baixas", use_container_width=True, type="primary"):
                    manual2 = load_manual()
                    manual2.setdefault("overrides", {})
                    manual2["overrides"][contrato_sel] = {}
                    for _, row in edited.iterrows():
                        num = str(int(row["Parcela"]))
                        manual2["overrides"][contrato_sel][num] = {
                            "Situação": row["Situação"],
                            "Emissão":  str(row["Emissão"]) if pd.notna(row["Emissão"]) else None,
                            "NF":       str(row["NF"]) if pd.notna(row["NF"]) else "",
                        }
                    save_manual(manual2)
                    st.cache_data.clear()
                    st.success("✅ Baixas salvas!")
                    st.rerun()

            with col_info2:
                fat = edited[edited["Situação"].isin(["FATURADO","ENTREGUE"])].shape[0]
                esp = edited[edited["Situação"] == "EM ESPERA"].shape[0]
                flt = edited[edited["Situação"] == "FALTA"].shape[0]
                st.caption(f"✅ {fat} faturadas/entregues · ⏳ {esp} em espera · ⚠️ {flt} com falta")

            # Progresso
            total_p = len(edited)
            pct_f   = (fat / total_p * 100) if total_p else 0
            cor_f   = "#24B78C" if pct_f >= 80 else ("#F5A623" if pct_f >= 40 else "#FF672F")
            st.markdown(
                f"<div style='background:#F5F0FA;border-radius:8px;padding:10px 16px;margin-top:8px'>"
                f"<div style='display:flex;justify-content:space-between;margin-bottom:5px'>"
                f"<span style='font-size:.72rem;color:#7E16B8;font-weight:600;text-transform:uppercase;"
                f"letter-spacing:1px'>Parcelas faturadas/entregues</span>"
                f"<span style='font-size:.8rem;font-weight:700;color:{cor_f}'>"
                f"{fat}/{total_p} · {pct_f:.0f}%</span></div>"
                f"<div style='background:#E0D4F0;border-radius:4px;height:8px'>"
                f"<div style='background:{cor_f};width:{min(pct_f,100):.0f}%;height:100%;border-radius:4px'>"
                f"</div></div></div>",
                unsafe_allow_html=True,
            )

# ── Tab Novo Contrato ──────────────────────────────────────────────────────────
with tab_novo:
    st.markdown("<div class='section-title'>Cadastrar Novo Contrato</div>", unsafe_allow_html=True)

    with st.form("form_novo_contrato", clear_on_submit=False):
        st.markdown("**📌 Dados do contrato**")
        fc1, fc2 = st.columns(2)
        nome_aba    = fc1.text_input("Nome do contrato *", placeholder="Ex: CLIENTE NOVO 2026")
        empresa_f   = fc1.selectbox("Empresa *", EMPRESAS)
        contratante = fc2.text_input("Contratante *", placeholder="Nome do cliente")
        data_pgto   = fc2.text_input("Dia de pagamento", placeholder="Ex: DIA 10")

        fc3, fc4 = st.columns(2)
        dt_assinatura = fc3.date_input("Data assinatura", value=date.today())
        dt_validade   = fc4.date_input("Validade", value=date.today().replace(year=date.today().year + 1))

        fc5, fc6 = st.columns(2)
        valor_total_f = fc5.number_input("Valor total do contrato (R$) *", min_value=0.0, step=1000.0, format="%.2f")
        n_parcelas_f  = fc6.number_input("Número de parcelas *", min_value=1, max_value=60, value=12, step=1)

        fc7, fc8 = st.columns(2)
        valor_parcela = fc7.number_input("Valor por parcela (R$)", min_value=0.0,
                                          value=round(valor_total_f / n_parcelas_f, 2) if n_parcelas_f else 0.0,
                                          step=100.0, format="%.2f")
        dt_primeira   = fc8.date_input("Data da 1ª parcela", value=date.today())

        fc9, fc10 = st.columns(2)
        kit      = fc9.text_input("Kit", placeholder="Ex: 4 KITS DETECÇÃO")
        amostras = fc10.text_input("Amostras", placeholder="Ex: 192 AMOSTRAS")
        obs      = st.text_area("Observações", height=70)

        submitted = st.form_submit_button("✅ Cadastrar contrato", type="primary", use_container_width=True)

        if submitted:
            erros = []
            if not nome_aba.strip():   erros.append("Nome do contrato é obrigatório.")
            if not contratante.strip():erros.append("Contratante é obrigatório.")
            if valor_total_f <= 0:     erros.append("Valor total deve ser maior que zero.")
            if nome_aba.strip() in detalhes_todos:
                erros.append(f"Já existe um contrato com o nome '{nome_aba.strip()}'.")

            if erros:
                for e in erros:
                    st.error(e)
            else:
                # Gera parcelas automaticamente
                parcelas_geradas = []
                saldo_restante   = valor_total_f
                for i in range(1, int(n_parcelas_f) + 1):
                    dt_parcela = dt_primeira.replace(
                        month=((dt_primeira.month - 1 + (i - 1)) % 12) + 1,
                        year=dt_primeira.year + ((dt_primeira.month - 1 + (i - 1)) // 12),
                    )
                    saldo_restante -= valor_parcela
                    parcelas_geradas.append({
                        "Parcela":  i,
                        "Emissão":  None,
                        "Valor":    valor_parcela,
                        "Saldo":    max(saldo_restante, 0),
                        "Situação": "EM ESPERA",
                        "NF":       "",
                    })

                novo = {
                    "info": {
                        "EMPRESA":     empresa_f,
                        "CONTRATANTE": contratante.strip(),
                        "ASSINATURA":  str(dt_assinatura),
                        "VALIDADE":    str(dt_validade),
                        "KIT":         kit.strip(),
                        "AMOSTRAS":    amostras.strip(),
                        "CONTRATO":    valor_total_f,
                        "PARCELA":     valor_parcela,
                        "SITUAÇÃO":    "EM VIGÊNCIA",
                        "DATA_PGTO":   data_pgto.strip(),
                        "CONSUMO":     0,
                        "SALDO":       valor_total_f,
                        "OBS":         obs.strip(),
                    },
                    "parcelas": parcelas_geradas,
                }

                manual2 = load_manual()
                manual2.setdefault("novos_contratos", {})
                manual2["novos_contratos"][nome_aba.strip()] = novo
                save_manual(manual2)
                st.cache_data.clear()
                st.success(f"✅ Contrato '{nome_aba.strip()}' cadastrado com {int(n_parcelas_f)} parcelas!")
                st.rerun()

    # Lista contratos manuais já cadastrados
    if novos:
        st.markdown("<div class='section-title'>Contratos Cadastrados Manualmente</div>",
                    unsafe_allow_html=True)
        for nome_c, dados_c in novos.items():
            with st.expander(f"📄 {nome_c}"):
                inf = dados_c["info"]
                col_x, col_y = st.columns([2, 1])
                with col_x:
                    st.markdown(f"**Empresa:** {inf.get('EMPRESA','—')}  |  "
                                f"**Contratante:** {inf.get('CONTRATANTE','—')}  |  "
                                f"**Valor:** {brl(inf.get('CONTRATO',0))}  |  "
                                f"**Validade:** {inf.get('VALIDADE','—')}")
                with col_y:
                    if st.button(f"🗑️ Excluir", key=f"del_{nome_c}"):
                        manual2 = load_manual()
                        manual2.get("novos_contratos", {}).pop(nome_c, None)
                        manual2.get("overrides", {}).pop(nome_c, None)
                        save_manual(manual2)
                        st.cache_data.clear()
                        st.rerun()

# ── Tab Encerrados ─────────────────────────────────────────────────────────────
with tab_enc_tab:
    if df_enc.empty:
        st.info("Nenhum contrato encerrado.")
    else:
        st.markdown("<div class='section-title'>Contratos Encerrados</div>", unsafe_allow_html=True)
        df_enc_show = df_enc.copy()
        df_enc_show["Contrato"] = df_enc_show["Contrato"].apply(brl)
        df_enc_show["Saldo"]    = df_enc_show["Saldo"].apply(brl)
        cols_enc = [c for c in ["Empresa","Contratante","Parcela","Contrato","Saldo"] if c in df_enc_show.columns]
        st.dataframe(df_enc_show[cols_enc], use_container_width=True, hide_index=True)
        st.caption(f"Valor histórico total: **{brl(df_enc['Contrato'].sum())}**")
