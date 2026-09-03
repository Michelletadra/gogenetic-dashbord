"""Página 6 — Sistema de Créditos de Clientes."""
import io
import streamlit as st
import pandas as pd
from collections import defaultdict
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers as xl_numbers)
from openpyxl.utils import get_column_letter

from utils import GLOBAL_CSS, brl, status_badge, sidebar_header, require_auth
from db_contratos import list_contratos as _list_contratos_all
from db_creditos import (
    list_clientes, insert_cliente, update_cliente, delete_cliente,
    list_notas, insert_nota, delete_nota,
    list_creditos, insert_credito, update_credito, delete_credito,
    list_movimentacoes, insert_movimentacao, delete_movimentacao,
)

st.set_page_config(page_title="Créditos | GoGenetic", page_icon="💳", layout="wide")
st.markdown(GLOBAL_CSS, unsafe_allow_html=True)
sidebar_header()
require_auth()

st.markdown("""
<p class="page-title">💳 Créditos de Clientes</p>
<p class="page-sub">Gestão completa de créditos · Grupo GoGenetic</p>
""", unsafe_allow_html=True)

# ── Carrega TUDO + constrói índices dentro do cache (executado UMA vez) ───────
@st.cache_data(ttl=600, show_spinner="⏳ Carregando créditos…")
def _load_all():
    """Carrega 4 tabelas em paralelo e pré-constrói todos os índices."""
    from concurrent.futures import ThreadPoolExecutor
    with ThreadPoolExecutor(max_workers=4) as ex:
        f_cli  = ex.submit(list_clientes)
        f_cred = ex.submit(list_creditos)
        f_nota = ex.submit(list_notas)
        f_movs = ex.submit(list_movimentacoes)
    clientes = f_cli.result()
    creditos = f_cred.result()
    notas    = f_nota.result()
    movs     = f_movs.result()

    # Saldo (valor_utilizado) é sempre CALCULADO a partir do histórico de
    # movimentações, nunca editado direto — soma UTILIZAÇÃO/USO/AJUSTE como
    # débito, ESTORNO como crédito de volta. "USO" é um tipo legado (import de
    # Excel antigo, ~125 registros) que significa a mesma coisa que
    # UTILIZAÇÃO — tem que entrar na soma senão o saldo calculado fica errado.
    # AJUSTE já vem com o sinal certo aplicado no lançamento (positivo = usa
    # mais saldo, negativo = devolve saldo).
    _debito_por_cred = defaultdict(float)
    _ultimo_uso_por_cred = {}
    for m in movs:
        cid = m.get("credito_id")
        valor = m.get("valor") or 0
        tipo = m.get("tipo")
        if tipo == "ESTORNO":
            _debito_por_cred[cid] -= valor
        else:  # UTILIZAÇÃO, USO (legado), AJUSTE
            _debito_por_cred[cid] += valor
        data_mov = m.get("data") or m.get("created_at")
        if data_mov and (cid not in _ultimo_uso_por_cred or data_mov > _ultimo_uso_por_cred[cid]):
            _ultimo_uso_por_cred[cid] = data_mov

    for c in creditos:
        c["_valor_utilizado_raw"] = c.get("valor_utilizado")
        c["valor_utilizado"]      = _debito_por_cred.get(c["id"], 0)
        c["ultimo_uso"]           = _ultimo_uso_por_cred.get(c["id"])

    # Índices — construídos aqui, reutilizados em todos os reruns
    cred_by_cli = defaultdict(list)
    for c in creditos:
        cred_by_cli[c["cliente_id"]].append(c)

    nota_by_cli = defaultdict(list)
    for n in notas:
        nota_by_cli[n["cliente_id"]].append(n)

    cred_map = {
        c["id"]: {"cliente_id": c["cliente_id"],
                  "cliente_nome": c.get("cliente_nome",""),
                  "valor_original": c.get("valor_original", 0)}
        for c in creditos
    }
    cred_to_cli = {cid: v["cliente_id"] for cid, v in cred_map.items()}

    # Enriquece movs com cliente_nome sem join no banco
    for m in movs:
        info = cred_map.get(m.get("credito_id"), {})
        if not m.get("cliente_nome"):
            m["cliente_nome"]   = info.get("cliente_nome", "")
        if "valor_original" not in m:
            m["valor_original"] = info.get("valor_original", 0)

    mov_by_cli = defaultdict(list)
    for m in movs:
        cli_id = cred_to_cli.get(m.get("credito_id"))
        if cli_id:
            mov_by_cli[cli_id].append(m)

    return {
        "clientes":    clientes,
        "creditos":    creditos,
        "notas":       notas,
        "movs":        movs,
        "cred_by_cli": cred_by_cli,
        "nota_by_cli": nota_by_cli,
        "mov_by_cli":  mov_by_cli,
        "cred_to_cli": cred_to_cli,
        "cred_map":    cred_map,
    }

@st.cache_data(ttl=600, show_spinner=False)
def _load_contratos():
    """Carrega contratos sob demanda (usado apenas no detalhe do cliente)."""
    contratos = _list_contratos_all()
    idx = defaultdict(list)
    for ct in contratos:
        if ct.get("cliente_id"):
            idx[ct["cliente_id"]].append(ct)
    return idx

def _clear_and_rerun():
    _load_all.clear()
    _load_contratos.clear()
    st.rerun()

def _tabs_persist(options: list, key: str) -> str:
    """Como st.tabs(), mas lembra qual aba estava selecionada entre reruns
    (st.tabs volta sempre pra primeira aba a cada rerun, o que derrubava
    o usuário de volta pro Painel no meio de um cadastro)."""
    last = st.session_state.get(key, options[0])
    if last not in options:
        last = options[0]
    sel = st.segmented_control(key, options, default=last, key=f"{key}_widget",
                                label_visibility="collapsed")
    if sel is None:
        sel = last
    st.session_state[key] = sel
    return sel

_top_l, _top_r = st.columns([5, 1])
with _top_r:
    if st.button("🔄 Atualizar dados", use_container_width=True,
                 help="Recarrega tudo do banco agora, sem esperar o cache (até 10min)."):
        _clear_and_rerun()

_data        = _load_all()
clientes_all = _data["clientes"]
creditos_all = _data["creditos"]
notas_all    = _data["notas"]
movs_all     = _data["movs"]
_cred_by_cli = _data["cred_by_cli"]
_nota_by_cli = _data["nota_by_cli"]
_mov_by_cli  = _data["mov_by_cli"]
_cred_to_cli = _data["cred_to_cli"]
_cred_map    = _data["cred_map"]

def _get_cont_by_cli():
    return _load_contratos()

def _nf_label(numero_nf) -> str:
    return f"NF {numero_nf}" if numero_nf else "Sem NF"

def _status_dot(status: str, dias: "int | None") -> str:
    """Bolinha de status — todo crédito tem uma, não só os VÁLIDO vencendo
    (antes só VÁLIDO ganhava bolinha e o resto ficava sem nada, inconsistente
    numa lista com status misturados)."""
    if status == "VÁLIDO":
        if dias is None:
            return "🟢"
        return "🔴" if dias <= 7 else ("🟡" if dias <= 30 else "🟢")
    if status == "EXPIRADO":
        return "🟠"
    if status == "UTILIZADO":
        return "⚪"
    return "⚫"  # CANCELADO ou outro

@st.cache_data(ttl=300, show_spinner="🔎 Buscando pedidos no eGestor…")
def _buscar_servicos_egestor(cliente_nome: str) -> list:
    """Busca serviços/pedidos REAIS do cliente nas 3 empresas eGestor (mesmo
    padrão de busca por substring de pages/7_Servicos.py — case-insensitive,
    sem normalização de acento), últimos 24 meses. Cada item ganha '_empresa'
    pra dar rastreabilidade (qual empresa eGestor o pedido pertence)."""
    from utils import get_clients
    termo = (cliente_nome or "").strip().lower()
    if not termo:
        return []
    dt_ini = (date.today() - timedelta(days=730)).isoformat()
    dt_fim = date.today().isoformat()
    achados = []
    for empresa, client in get_clients().items():
        try:
            servicos = client.get_servicos(dt_ini, dt_fim)
        except Exception:
            continue
        for s in servicos:
            nome_contato = (s.get("nomeContato") or "").lower()
            if termo in nome_contato or (nome_contato and nome_contato in termo):
                s = dict(s)
                s["_empresa"] = empresa
                achados.append(s)
    achados.sort(key=lambda s: s.get("dtVenda") or "", reverse=True)
    return achados

def _render_form_consumo(cr, key_suffix=""):
    """Formulário de registrar consumo — usado tanto na aba 💳 Créditos
    (Lista/Registrar Consumo) quanto no perfil do cliente (🧑‍🤝‍🧑 Clientes),
    pra não ter duas versões divergentes da mesma coisa.

    O serviço/pedido é sempre um registro REAL do eGestor (rastreabilidade
    NF → Crédito → Consumo → Serviço) — nada de descrição digitada à mão."""
    saldo_d = (cr["valor_original"] or 0) - (cr["valor_utilizado"] or 0)
    st.markdown(f"**Saldo disponível: {brl(saldo_d)}**")

    servicos_disp = _buscar_servicos_egestor(cr.get("cliente_nome"))
    if not servicos_disp:
        st.warning(
            f"⚠️ Nenhum pedido/serviço encontrado no eGestor pra "
            f"**{cr.get('cliente_nome','este cliente')}** nos últimos 24 meses. "
            f"Cadastre o pedido no eGestor antes de registrar o consumo aqui."
        )
        return

    serv_opts = {
        f"#{s.get('codigo','?')} · {s.get('_empresa','')} · "
        f"{s.get('dtVenda','—')} · {brl(s.get('valorTotal'))} · {s.get('situacaoOS') or s.get('situacao') or ''}": s
        for s in servicos_disp
    }
    serv_label = st.selectbox(
        "Serviço/Pedido * (real, do eGestor)", list(serv_opts.keys()),
        key=f"serv_sel_{cr['id']}{key_suffix}",
    )
    serv_sel = serv_opts[serv_label]

    with st.form(f"form_consumo_dash_{cr['id']}{key_suffix}", clear_on_submit=True):
        v_uso, _ = _valor_input(
            "Valor consumido (R$) *", key=f"v_uso_{cr['id']}{key_suffix}",
            valor_atual=float(serv_sel.get("valorTotal") or 0) or None,
            help=f"Saldo disponível: {brl(saldo_d)}. Pode passar — o excedente vira saldo "
                 f"negativo, cobrado do cliente à parte.",
        )

        cc, cd = st.columns(2)
        data_serv = cc.date_input("Data do serviço", value=date.today(),
                                   key=f"data_serv_{cr['id']}{key_suffix}")
        resp      = cd.text_input("Responsável", key=f"resp_{cr['id']}{key_suffix}")
        obs_u = st.text_area("Observação (opcional)", height=60, key=f"obs_u_{cr['id']}{key_suffix}")

        if st.form_submit_button("💸 Registrar Consumo", use_container_width=True):
            if not v_uso or v_uso <= 0:
                st.error("❌ Informe um valor válido de consumo.")
            else:
                # Saldo é sempre CALCULADO a partir das movimentações (ver
                # _load_all) — aqui só grava o lançamento, nunca valor_utilizado.
                novo_saldo = saldo_d - v_uso
                if novo_saldo <= 0 and cr["status"] != "UTILIZADO":
                    update_credito(cr["id"], {"status": "UTILIZADO"})
                insert_movimentacao({
                    "credito_id":        cr["id"],
                    "tipo":              "UTILIZAÇÃO",
                    "valor":             float(v_uso),
                    "data":              str(data_serv),
                    "responsavel":       resp or None,
                    "observacao":        obs_u or None,
                    "descricao_servico": f"Pedido #{serv_sel.get('codigo','?')} ({serv_sel.get('_empresa','')})",
                    "codigo_servico":    str(serv_sel.get("codigo") or "") or None,
                    "servico_empresa":   serv_sel.get("_empresa"),
                })
                if novo_saldo < 0:
                    st.session_state["_consumo_ok"] = (
                        f"✅ Consumo de {brl(v_uso)} registrado! "
                        f"Saldo ficou negativo em {brl(abs(novo_saldo))} — cobrar esse valor do cliente."
                    )
                else:
                    st.session_state["_consumo_ok"] = f"✅ Consumo de {brl(v_uso)} registrado!"
                _clear_and_rerun()

@st.dialog("💸 Registrar consumo")
def _abrir_dialog_consumo(cr):
    """Aberto sob demanda (não a cada linha da lista) — antes o formulário
    inteiro de cada crédito era montado em popover pra TODAS as linhas em
    todo carregamento de página, o que deixava a Lista lenta com muitos
    créditos (100+). Com st.dialog só constrói quando alguém clica."""
    alerta_exp = " ⚠️ EXPIRADO" if cr["status"] == "EXPIRADO" else ""
    st.caption(f"{cr.get('cliente_nome','?')} — {_nf_label(cr.get('numero_nf'))}{alerta_exp}")
    _render_form_consumo(cr, key_suffix=f"_dlg_{cr['id']}")

@st.dialog("🗑️ Excluir crédito")
def _abrir_dialog_excluir(cr):
    st.caption(f"Excluir crédito de {cr.get('cliente_nome','?')} — {brl(cr.get('valor_original'))}")
    movs_do_credito = [m for m in movs_all if m.get("credito_id") == cr["id"]]
    if movs_do_credito:
        st.warning(
            f"⚠️ Tem {len(movs_do_credito)} movimentação(ões) registrada(s) — "
            f"excluir apaga esse histórico junto."
        )
    confirma_del = st.checkbox("Confirmo a exclusão", key=f"del_confirm_dlg_{cr['id']}")
    if st.button("🗑️ Excluir definitivamente", key=f"del_btn_dlg_{cr['id']}",
                 disabled=not confirma_del, use_container_width=True):
        delete_credito(cr["id"])
        st.session_state["_edit_cred_ok"] = f"🗑️ Crédito de {cr.get('cliente_nome','?')} excluído."
        _clear_and_rerun()

def _attach_nf_control(cr, key_suffix=""):
    """Campo rápido pra anexar (ou trocar) a NF de um crédito, sem precisar
    abrir outra aba. Digitou um número novo -> cria a NF e já vincula."""
    with st.form(f"form_nf_{cr['id']}{key_suffix}", clear_on_submit=True):
        nf1, nf2 = st.columns([3, 1])
        nf_num = nf1.text_input(
            "Número da NF", value=cr.get("numero_nf") or "",
            key=f"nf_num_{cr['id']}{key_suffix}", placeholder="ex: 6640",
        )
        salvar_nf = nf2.form_submit_button("💾 Salvar NF", use_container_width=True)
        if salvar_nf:
            nf_num = nf_num.strip()
            if not nf_num:
                st.error("❌ Digite o número da NF antes de salvar.")
            else:
                nf_id = insert_nota({
                    "numero_nf":    nf_num,
                    "cliente_id":   cr["cliente_id"],
                    "data_emissao": str(date.today()),
                    "valor_total":  float(cr.get("valor_original") or 0),
                })
                update_credito(cr["id"], {"nota_fiscal_id": nf_id})
                st.session_state["_edit_cred_ok"] = f"✅ NF {nf_num} vinculada ao crédito!"
                _clear_and_rerun()

def _parse_valor(texto: str):
    """Converte texto digitado em R$ pra float — aceita '2500', '2500.00',
    '2500,00' e '2.500,00' (o campo numérico nativo do navegador só aceita
    ponto e rejeita vírgula, o que fazia o valor digitado sumir/dar errado)."""
    if not texto:
        return None
    t = texto.strip().replace("R$", "").replace(" ", "")
    if not t:
        return None
    if "," in t:
        t = t.replace(".", "").replace(",", ".")
    try:
        val = float(t)
    except ValueError:
        return None
    return val if val >= 0 else None

def _valor_input(label: str, key: str, valor_atual: float = None, help: str = None) -> tuple:
    """Campo de texto pra valores em R$ — mostra o valor interpretado embaixo
    pra confirmar antes de salvar. Retorna (valor_float_ou_None, texto_digitado)."""
    default = f"{valor_atual:.2f}".replace(".", ",") if valor_atual is not None else ""
    texto = st.text_input(label, value=default, key=key,
                           placeholder="ex: 2500,00", help=help)
    val = _parse_valor(texto)
    if texto and val is None:
        st.caption("⚠️ Não entendi esse valor — use só números, vírgula e ponto.")
    elif val is not None:
        st.caption(f"= {brl(val)}")
    return val, texto

def _resumo_mem(cli_id: int) -> dict:
    """Resumo do cliente — lê direto dos índices já em cache."""
    creds  = _cred_by_cli.get(cli_id, [])
    notas_ = _nota_by_cli.get(cli_id, [])
    valid  = [c for c in creds if c["status"] == "VÁLIDO"]
    expir  = [c for c in creds if c["status"] == "EXPIRADO"]
    saldo  = lambda lst: sum((c.get("valor_original") or 0) - (c.get("valor_utilizado") or 0) for c in lst)
    return {
        "qtd_validos":     len(valid),
        "qtd_expirados":   len(expir),
        "saldo_valido":    saldo(valid),
        "saldo_expirado":  saldo(expir),
        "total_utilizado": sum(c.get("valor_utilizado") or 0 for c in creds),
        "qtd_notas":       len(notas_),
    }

def _saldo_credito(cr) -> float:
    return (cr.get("valor_original") or 0) - (cr.get("valor_utilizado") or 0)

def _divergencia_credito(cr) -> float:
    """Diferença entre o valor_utilizado gravado no banco (campo legado —
    editado direto no passado, pode estar errado) e o calculado a partir das
    movimentações reais de hoje (sempre correto — ver _load_all). != 0 quer
    dizer que esse crédito precisa de um ⚖️ Ajuste documentando a correção."""
    bruto = cr.get("_valor_utilizado_raw")
    if bruto is None:
        return 0.0
    calc = cr.get("valor_utilizado") or 0
    return round((bruto or 0) - calc, 2)

def _creditos_ja_ajustados() -> set:
    return {m.get("credito_id") for m in movs_all if m.get("tipo") == "AJUSTE"}

def _credito_precisa_conciliar(cr, ja_ajustados: set = None) -> bool:
    """True quando há divergência entre o campo legado e o extrato real E
    ainda ninguém documentou isso com um Ajuste — depois que alguém registra
    o Ajuste (mesmo de valor 0, só pra explicar), o crédito para de ser
    sinalizado, mesmo que o campo legado nunca mude (ele não é mais lido pra
    nada, só serve de rastro histórico)."""
    if abs(_divergencia_credito(cr)) < 0.01:
        return False
    if ja_ajustados is None:
        ja_ajustados = _creditos_ja_ajustados()
    return cr["id"] not in ja_ajustados

def _creditos_divergentes() -> list:
    ja_ajustados = _creditos_ja_ajustados()
    return [c for c in creditos_all if _credito_precisa_conciliar(c, ja_ajustados)]

def _parse_valor_signed(texto: str):
    """Como _parse_valor, mas aceita '-' na frente pra ajustes que devolvem
    saldo (crédito de volta pro cliente)."""
    if not texto:
        return None
    t = texto.strip()
    neg = t.startswith("-")
    if neg:
        t = t[1:].strip()
    val = _parse_valor(t)
    if val is None:
        return None
    return -val if neg else val

@st.dialog("⚖️ Ajuste de saldo")
def _abrir_dialog_ajuste(cr):
    """Único jeito de corrigir o saldo calculado sem editar valor_utilizado
    direto (a causa da divergência original) — todo ajuste vira uma
    movimentação tipo AJUSTE, com motivo OBRIGATÓRIO, sempre visível no
    extrato. Valor positivo consome mais saldo (débito); negativo devolve
    saldo (estorna um débito indevido)."""
    st.caption(f"{cr.get('cliente_nome', '?')} — {_nf_label(cr.get('numero_nf'))} · "
               f"saldo atual: {brl(_saldo_credito(cr))}")
    div = _divergencia_credito(cr)
    if abs(div) >= 0.01:
        st.warning(
            f"⚠️ O campo antigo do banco registra {brl(cr.get('_valor_utilizado_raw') or 0)} "
            f"utilizados, mas a soma real das movimentações mostra {brl(cr.get('valor_utilizado') or 0)} "
            f"— diferença de {brl(div)}. Confirme com o histórico do cliente se realmente houve "
            f"esse consumo antes de decidir o valor do ajuste abaixo."
        )
    with st.form(f"form_ajuste_{cr['id']}"):
        valor_txt = st.text_input(
            "Valor do ajuste (R$) *", key=f"ajuste_valor_{cr['id']}",
            placeholder="ex: 1500,00 ou -1500,00",
            help="Positivo = usa mais saldo (débito). Negativo = devolve saldo (estorna um "
                 "débito indevido). O saldo do extrato já é sempre recalculado — isso só "
                 "registra o lançamento com o motivo.",
        )
        resp   = st.text_input("Responsável", key=f"ajuste_resp_{cr['id']}")
        motivo = st.text_area(
            "Motivo do ajuste * (obrigatório)", height=90, key=f"ajuste_motivo_{cr['id']}",
            placeholder="Explique a divergência encontrada e a correção aplicada — fica "
                        "visível no extrato pra sempre.",
        )
        if st.form_submit_button("⚖️ Registrar ajuste", use_container_width=True):
            val = _parse_valor_signed(valor_txt)
            if not valor_txt.strip() or val is None:
                st.error("❌ Informe um valor de ajuste válido (pode ser negativo).")
            elif not motivo.strip():
                st.error("❌ O motivo é obrigatório — todo ajuste precisa de uma explicação registrada.")
            else:
                insert_movimentacao({
                    "credito_id": cr["id"],
                    "tipo":       "AJUSTE",
                    "valor":      float(val),
                    "data":       str(date.today()),
                    "responsavel": resp or None,
                    "observacao": motivo.strip(),
                })
                st.session_state["_ajuste_ok"] = f"⚖️ Ajuste registrado no extrato de {cr.get('cliente_nome', '?')}."
                _clear_and_rerun()

@st.dialog("➕ Adicionar crédito")
def _abrir_dialog_add_credito(cli):
    st.caption(f"Novo crédito para {cli['nome']}")
    notas_cl = _nota_by_cli.get(cli["id"], [])
    nf_opts = {"— Sem NF —": None}
    nf_opts.update({f"NF {n['numero_nf']}": n["id"] for n in notas_cl})

    with st.form(f"form_add_cred_{cli['id']}", clear_on_submit=True):
        valor, _ = _valor_input("Valor (R$) *", key=f"add_cred_valor_{cli['id']}",
                                 help="Valor real do crédito concedido — confira antes de cadastrar.")
        venc = st.date_input("Vencimento *", value=date.today() + timedelta(days=30),
                              key=f"add_cred_venc_{cli['id']}")
        c1, c2 = st.columns(2)
        nf_sel  = c1.selectbox("NF já cadastrada", list(nf_opts.keys()), key=f"add_cred_nfsel_{cli['id']}")
        nf_nova = c2.text_input("Ou NF nova", placeholder="ex: 6640", key=f"add_cred_nfnova_{cli['id']}")

        contratos_cl = [ct for ct in _get_cont_by_cli().get(cli["id"], [])
                         if ct.get("status_real") not in ("ENCERRADO", "RESCINDIDO")]
        ct_opts = {"— Sem contrato —": None}
        ct_opts.update({
            f"{ct.get('contratante', '?')} ({ct.get('empresa_gg', '—')}) · {ct.get('tipo_contrato', '—')}": ct.get("id")
            for ct in contratos_cl
        })
        ct_sel = st.selectbox("Contrato vinculado (opcional)", list(ct_opts.keys()), key=f"add_cred_ct_{cli['id']}")
        obs = st.text_area("Observações", height=60, key=f"add_cred_obs_{cli['id']}")

        if st.form_submit_button("➕ Adicionar crédito", use_container_width=True):
            if valor is None:
                st.error("❌ Informe um valor válido.")
            elif valor < 1:
                st.error(f"❌ Valor muito baixo ({brl(valor)}). Confirme se digitou certo antes de cadastrar.")
            elif nf_nova.strip() and nf_sel != "— Sem NF —":
                st.error("❌ Escolha uma NF já cadastrada OU digite uma nova — não os dois.")
            else:
                nota_fiscal_id_sel = nf_opts[nf_sel]
                if nf_nova.strip():
                    nota_fiscal_id_sel = insert_nota({
                        "numero_nf": nf_nova.strip(), "cliente_id": cli["id"],
                        "data_emissao": str(date.today()), "valor_total": float(valor),
                    })
                payload = {
                    "cliente_id": cli["id"], "nota_fiscal_id": nota_fiscal_id_sel,
                    "valor_original": float(valor), "data_vencimento": str(venc),
                    "observacoes": obs or None, "contrato_id": ct_opts[ct_sel],
                }
                sucesso = False
                try:
                    insert_credito(payload)
                    sucesso = True
                except Exception as e:
                    if "contrato_id" in str(e):
                        try:
                            payload.pop("contrato_id")
                            insert_credito(payload)
                            sucesso = True
                        except Exception as e2:
                            st.error(f"❌ Não foi possível cadastrar o crédito: {e2}")
                    else:
                        st.error(f"❌ Não foi possível cadastrar o crédito: {e}")
                if sucesso:
                    st.session_state["_wallet_ok"] = f"✅ Crédito de {brl(valor)} adicionado para {cli['nome']}!"
                    _clear_and_rerun()

@st.dialog("💸 Usar crédito")
def _abrir_dialog_usar(cli, creds_disponiveis):
    if len(creds_disponiveis) == 1:
        cr = creds_disponiveis[0]
    else:
        opts = {
            f"{_nf_label(c.get('numero_nf'))} — saldo {brl(_saldo_credito(c))}"
            f"{' ⚠️ EXPIRADO' if c['status'] == 'EXPIRADO' else ''} (#{c['id']})": c
            for c in creds_disponiveis
        }
        label = st.selectbox("Qual crédito usar?", list(opts.keys()), key=f"usar_sel_{cli['id']}")
        cr = opts[label]
    alerta_exp = " ⚠️ EXPIRADO" if cr["status"] == "EXPIRADO" else ""
    st.caption(f"{_nf_label(cr.get('numero_nf'))}{alerta_exp}")
    _render_form_consumo(cr, key_suffix=f"_wallet_{cr['id']}")

@st.dialog("✏️ Editar crédito")
def _abrir_dialog_editar_credito(cr):
    st.caption(f"{cr.get('cliente_nome', '?')} — {_nf_label(cr.get('numero_nf'))}")
    notas_cl_edit = _nota_by_cli.get(cr.get("cliente_id"), [])
    nf_opts_edit = {"— Sem NF —": None}
    nf_opts_edit.update({f"NF {n['numero_nf']}": n["id"] for n in notas_cl_edit})
    nf_labels_edit = list(nf_opts_edit.keys())
    nf_atual_idx = next((i for i, v in enumerate(nf_opts_edit.values()) if v == cr.get("nota_fiscal_id")), 0)

    contratos_cl_edit = _get_cont_by_cli().get(cr.get("cliente_id"), [])
    ct_opts_edit = {"— Sem contrato —": None}
    ct_opts_edit.update({
        f"{ct.get('contratante', '?')} ({ct.get('empresa_gg', '—')}) · {ct.get('tipo_contrato', '—')}": ct.get("id")
        for ct in contratos_cl_edit
    })
    ct_labels_edit = list(ct_opts_edit.keys())
    ct_atual_idx = next((i for i, v in enumerate(ct_opts_edit.values()) if v == cr.get("contrato_id")), 0)

    with st.form(f"form_edit_cred_wallet_{cr['id']}"):
        novo_valor, _ = _valor_input("Valor original (R$)", key=f"edit_valor_w_{cr['id']}",
                                      valor_atual=float(cr.get("valor_original") or 0))
        venc_atual = pd.to_datetime(cr.get("data_vencimento"), errors="coerce")
        novo_venc = st.date_input(
            "Vencimento", key=f"edit_venc_w_{cr['id']}",
            value=venc_atual.date() if pd.notna(venc_atual) else date.today() + timedelta(days=30),
        )
        status_opts = ["VÁLIDO", "EXPIRADO", "UTILIZADO", "CANCELADO"]
        novo_status = st.selectbox(
            "Status", status_opts, key=f"edit_status_w_{cr['id']}",
            index=status_opts.index(cr.get("status")) if cr.get("status") in status_opts else 0,
        )
        novo_nf_label = st.selectbox("NF vinculada", nf_labels_edit, index=nf_atual_idx, key=f"edit_nf_w_{cr['id']}")
        novo_ct_label = st.selectbox("Contrato vinculado", ct_labels_edit, index=ct_atual_idx, key=f"edit_ct_w_{cr['id']}")
        nova_obs = st.text_area("Observações", key=f"edit_obs_w_{cr['id']}",
                                 value=cr.get("observacoes") or "", height=80)
        if st.form_submit_button("💾 Salvar alterações", use_container_width=True):
            if novo_valor is None:
                st.error("❌ Informe um valor válido antes de salvar.")
            else:
                update_credito(cr["id"], {
                    "valor_original":  float(novo_valor),
                    "data_vencimento": str(novo_venc),
                    "status":          novo_status,
                    "nota_fiscal_id":  nf_opts_edit[novo_nf_label],
                    "contrato_id":     ct_opts_edit[novo_ct_label],
                    "observacoes":     nova_obs or None,
                })
                st.session_state["_wallet_ok"] = f"✅ Crédito de {cr.get('cliente_nome', '?')} atualizado!"
                _clear_and_rerun()

@st.dialog("🗑️ Apagar movimentação")
def _abrir_dialog_del_mov(m):
    st.caption(f"{m.get('tipo')} — {brl(m.get('valor'))} — {m.get('data') or ''}")
    if m.get("observacao"):
        st.caption(f"Obs: {m['observacao']}")
    st.warning("Isso também desfaz o efeito no saldo do crédito (o extrato é recalculado na hora).")
    if st.button("🗑️ Confirmar exclusão", key=f"del_mov_confirm_{m['id']}", use_container_width=True):
        # Saldo é calculado a partir das movimentações (_load_all) — apagar
        # já corrige o saldo sozinho no próximo load. Só falta destravar o
        # status se o crédito não estiver mais 100% usado.
        cred = next((c for c in creditos_all if c["id"] == m.get("credito_id")), None)
        if (cred and m.get("tipo") in ("UTILIZAÇÃO", "USO") and cred.get("status") == "UTILIZADO"):
            novo_ut = max(0, (cred.get("valor_utilizado") or 0) - (m.get("valor") or 0))
            if novo_ut < (cred.get("valor_original") or 0):
                update_credito(cred["id"], {"status": "VÁLIDO"})
        delete_movimentacao(m["id"])
        st.session_state["_wallet_ok"] = "✅ Movimentação apagada e saldo recalculado."
        _clear_and_rerun()

# ── Tabs principais ───────────────────────────────────────────────────────────
main_tab = _tabs_persist(
    ["👛 Carteiras", "📑 Relatório Mensal"],
    key="cred_main_tab",
)

# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 — CARTEIRAS
# ══════════════════════════════════════════════════════════════════════════════
if main_tab == "👛 Carteiras":
    for _flash_key in ("_wallet_ok", "_ajuste_ok", "_consumo_ok"):
        if st.session_state.get(_flash_key):
            st.success(st.session_state.pop(_flash_key))

    divergentes = _creditos_divergentes()
    if divergentes:
        nomes_div = sorted({c.get("cliente_nome", "?") for c in divergentes})
        st.markdown(f"""
        <div style='background:#FCEFD9;border-left:4px solid #B4720A;border-radius:8px;
                    padding:12px 16px;margin-bottom:18px;font-size:.86rem;color:#4B3B1E'>
          <b>⚠️ {len(divergentes)} crédito(s) com saldo divergente</b> — o valor gravado
          antigamente no banco não bate com a soma real das movimentações de hoje.
          O saldo mostrado abaixo já está correto (sempre calculado do extrato) — falta
          só registrar um ⚖️ Ajuste em cada um documentando a correção, pra fechar a
          divergência de vez: {", ".join(nomes_div)}.
        </div>""", unsafe_allow_html=True)

    col1, col2 = st.columns([3, 1])
    busca = col1.text_input("🔎 Buscar cliente ou NF")
    if col2.button("➕ Novo cliente", use_container_width=True):
        st.session_state["_wallet_novo_cli"] = True

    if st.session_state.get("_wallet_novo_cli"):
        with st.form("form_novo_cli_wallet", clear_on_submit=True):
            st.markdown("**Novo cliente**")
            c1, c2 = st.columns(2)
            nome  = c1.text_input("Nome *")
            email = c2.text_input("Email")
            obs   = st.text_area("Observações", height=60)
            s1, s2 = st.columns(2)
            if s1.form_submit_button("✅ Salvar", use_container_width=True):
                if nome.strip():
                    insert_cliente({"nome": nome.strip(), "email": email or None, "observacoes": obs or None})
                    st.session_state["_wallet_novo_cli"] = False
                    st.session_state["wallet_cli_sel"] = nome.strip()
                    st.session_state["_wallet_ok"] = f"✅ {nome} cadastrado!"
                    _clear_and_rerun()
                else:
                    st.error("Nome é obrigatório.")
            if s2.form_submit_button("❌ Cancelar", use_container_width=True):
                st.session_state["_wallet_novo_cli"] = False
                st.rerun()

    # Busca por nome de cliente OU por número de NF — pula direto pro dono da NF.
    termo = (busca or "").strip().lower()
    if termo:
        clientes_match  = {c["id"] for c in clientes_all if termo in c["nome"].lower()}
        nf_match_cli_ids = {c["cliente_id"] for c in creditos_all
                             if termo in str(c.get("numero_nf") or "").lower()}
        ids_match = clientes_match | nf_match_cli_ids
        lista = [c for c in clientes_all if c["id"] in ids_match]
    else:
        lista = clientes_all

    if not lista:
        st.info("Nenhum cliente encontrado.")
    else:
        # ── Tabela resumo — mesma leveza de antes, ordenada por saldo válido ──
        rows_tab = []
        for cli_r in lista:
            res_r = _resumo_mem(cli_r["id"])
            rows_tab.append({
                "Cliente":      cli_r["nome"],
                "Saldo Válido": res_r.get("saldo_valido", 0),
                "Créditos ✅":  res_r.get("qtd_validos", 0),
                "Créditos ❌":  res_r.get("qtd_expirados", 0),
                "Utilizado":    res_r.get("total_utilizado", 0),
            })
        df_clientes = pd.DataFrame(rows_tab).sort_values("Saldo Válido", ascending=False)
        df_show_cli = df_clientes.copy()
        df_show_cli["Saldo Válido"] = df_show_cli["Saldo Válido"].apply(brl)
        df_show_cli["Utilizado"]    = df_show_cli["Utilizado"].apply(brl)
        st.dataframe(df_show_cli, use_container_width=True, hide_index=True)

        st.markdown("---")

        nomes_lista = [c["nome"] for c in lista]
        if st.session_state.get("wallet_cli_sel") not in nomes_lista:
            st.session_state["wallet_cli_sel"] = nomes_lista[0]
        cli_sel_nome = st.selectbox("👛 Ver carteira de:", nomes_lista, key="wallet_cli_sel")
        cli = next(c for c in lista if c["nome"] == cli_sel_nome)

        creds_cli = _cred_by_cli.get(cli["id"], [])
        movs_cli  = _mov_by_cli.get(cli["id"], [])
        res       = _resumo_mem(cli["id"])
        creds_com_saldo = [c for c in creds_cli if c["status"] != "CANCELADO" and _saldo_credito(c) > 0]

        # ── Cabeçalho da carteira ────────────────────────────────────────────
        hcol1, hcol2 = st.columns([2.2, 1])
        with hcol1:
            st.markdown(f"### {cli['nome']}")
            sub_bits = [f"{res.get('qtd_validos', 0)} crédito(s) válido(s)"]
            if res.get("qtd_expirados", 0):
                sub_bits.append(f"{res['qtd_expirados']} expirado(s)")
            st.caption(" · ".join(sub_bits))
        with hcol2:
            extra_exp = ""
            if res.get("saldo_expirado", 0) > 0.009:
                extra_exp = (f"<div style='font-size:.78rem;color:#B4720A'>"
                             f"+ {brl(res.get('saldo_expirado', 0))} em créditos expirados</div>")
            st.markdown(f"""
            <div style='text-align:right'>
              <div style='font-size:.7rem;text-transform:uppercase;letter-spacing:1px;color:#8B6BAE'>
                Saldo disponível</div>
              <div style='font-size:1.6rem;font-weight:800;color:#159A73'>{brl(res.get('saldo_valido', 0))}</div>
              {extra_exp}
            </div>""", unsafe_allow_html=True)

        bcol1, bcol2, bcol3, bcol4 = st.columns(4)
        if bcol1.button("➕ Adicionar crédito", use_container_width=True, key=f"btn_add_{cli['id']}"):
            _abrir_dialog_add_credito(cli)
        if bcol2.button("💸 Usar crédito", use_container_width=True,
                         disabled=not creds_com_saldo, key=f"btn_use_{cli['id']}"):
            _abrir_dialog_usar(cli, creds_com_saldo)
        if bcol3.button("⚖️ Ajuste", use_container_width=True,
                         disabled=not creds_cli, key=f"btn_adj_{cli['id']}"):
            if len(creds_cli) == 1:
                _abrir_dialog_ajuste(creds_cli[0])
            else:
                st.session_state["_wallet_escolher_ajuste"] = cli["id"]
        if bcol4.button("📄 Nova NF", use_container_width=True, key=f"btn_nf_{cli['id']}"):
            st.session_state[f"_wallet_nova_nf_{cli['id']}"] = True

        if st.session_state.get("_wallet_escolher_ajuste") == cli["id"]:
            opts_adj = {f"{_nf_label(c.get('numero_nf'))} — saldo {brl(_saldo_credito(c))} (#{c['id']})": c
                        for c in creds_cli}
            sel_adj = st.selectbox("Ajuste em qual crédito?", list(opts_adj.keys()), key=f"adj_pick_{cli['id']}")
            if st.button("Continuar", key=f"adj_pick_go_{cli['id']}"):
                st.session_state.pop("_wallet_escolher_ajuste")
                _abrir_dialog_ajuste(opts_adj[sel_adj])

        if st.session_state.get(f"_wallet_nova_nf_{cli['id']}"):
            with st.form(f"nova_nf_wallet_{cli['id']}", clear_on_submit=True):
                st.caption("Nova NF (sem crédito automático — use ➕ Adicionar crédito pra isso)")
                n1, n2 = st.columns(2)
                num_nf = n1.text_input("Número NF")
                with n2:
                    valor_nf, _ = _valor_input("Valor (R$)", key=f"valor_nf_w_{cli['id']}")
                data_em = st.date_input("Data emissão")
                sb1, sb2 = st.columns(2)
                if sb1.form_submit_button("➕ Cadastrar NF", use_container_width=True):
                    if not num_nf.strip():
                        st.error("❌ Informe o número da NF.")
                    else:
                        insert_nota({"numero_nf": num_nf.strip(), "cliente_id": cli["id"],
                                     "data_emissao": str(data_em), "valor_total": float(valor_nf or 0)})
                        st.session_state[f"_wallet_nova_nf_{cli['id']}"] = False
                        st.session_state["_wallet_ok"] = f"✅ NF {num_nf} cadastrada!"
                        _clear_and_rerun()
                if sb2.form_submit_button("❌ Cancelar", use_container_width=True):
                    st.session_state[f"_wallet_nova_nf_{cli['id']}"] = False
                    st.rerun()

        # ── Créditos (envelopes) abertos ─────────────────────────────────────
        if creds_cli:
            st.markdown("<br>", unsafe_allow_html=True)
            hoje_ts = pd.Timestamp.today().normalize()
            creds_sorted = sorted(creds_cli, key=lambda c: c.get("data_vencimento") or "")
            for start in range(0, len(creds_sorted), 4):
                row_cols = st.columns(4)
                for j, cr in enumerate(creds_sorted[start:start + 4]):
                    saldo_cr = _saldo_credito(cr)
                    venc = pd.to_datetime(cr.get("data_vencimento"), errors="coerce")
                    dias = int((venc - hoje_ts).days) if pd.notna(venc) else None
                    dot  = _status_dot(cr["status"], dias)
                    tag_div = (" <span style='color:#B4720A;font-size:.66rem;font-weight:700'>"
                               "⚠️ divergente</span>") if _credito_precisa_conciliar(cr) else ""
                    with row_cols[j]:
                        st.markdown(f"""
                        <div style='background:#fff;border:1px solid #EAE6F4;border-radius:12px;
                                    padding:12px 14px;margin-bottom:8px'>
                          <div style='font-size:.82rem;font-weight:600'>{dot} {_nf_label(cr.get('numero_nf'))}{tag_div}</div>
                          <div style='font-family:monospace;font-size:1.05rem;font-weight:700;margin-top:3px'>
                            {brl(saldo_cr)}</div>
                          <div style='font-size:.72rem;color:#8B6BAE'>{cr['status']} · de {brl(cr.get('valor_original'))}
                            {f" · vence {venc.strftime('%d/%m/%Y')}" if pd.notna(venc) else ""}</div>
                        </div>""", unsafe_allow_html=True)
                        ec1, ec2 = st.columns(2)
                        if ec1.button("✏️ Editar", key=f"btn_edit_{cr['id']}", use_container_width=True):
                            _abrir_dialog_editar_credito(cr)
                        if ec2.button("🗑️ Excluir", key=f"btn_delc_{cr['id']}", use_container_width=True):
                            _abrir_dialog_excluir(cr)

        # ── Extrato ───────────────────────────────────────────────────────────
        st.markdown("#### 📜 Extrato")

        cred_id_to_nf = {cr["id"]: _nf_label(cr.get("numero_nf")) for cr in creds_cli}
        linhas = []
        for cr in creds_cli:
            linhas.append({
                "_ord": cr.get("created_at") or cr.get("data_vencimento") or "",
                "data": cr.get("created_at", ""),
                "tipo": "ENTRADA",
                "desc": f"Crédito concedido — {_nf_label(cr.get('numero_nf'))}",
                "sub":  cr.get("observacoes") or "",
                "valor": cr.get("valor_original") or 0,
                "mov":  None,
            })
        for m in movs_cli:
            tipo  = m.get("tipo")
            raw   = m.get("valor") or 0
            # Mesma lógica de sinal do cálculo de saldo em _load_all, invertida:
            # o que lá é débito (reduz saldo) aqui aparece negativo no extrato;
            # o que lá é crédito de volta aparece positivo.
            valor_disp = raw if tipo == "ESTORNO" else -raw
            desc_tipo = {"AJUSTE": "Ajuste de saldo", "ESTORNO": "Estorno"}.get(tipo)
            desc = m.get("descricao_servico") or desc_tipo or tipo
            nf_tag = cred_id_to_nf.get(m.get("credito_id"), "")
            linhas.append({
                "_ord": m.get("data") or m.get("created_at") or "",
                "data": m.get("data") or "",
                "tipo": tipo,
                "desc": f"{desc} — {nf_tag}" if nf_tag else desc,
                "sub":  m.get("observacao") or "",
                "valor": valor_disp,
                "mov":  m,
            })
        linhas.sort(key=lambda l: l["_ord"], reverse=True)

        if not linhas:
            st.info("Sem movimentações ainda.")
        else:
            buf_extrato = io.BytesIO()
            df_extrato = pd.DataFrame([{
                "Data": l["data"], "Tipo": l["tipo"], "Descrição": l["desc"],
                "Observação": l["sub"], "Valor": l["valor"],
            } for l in linhas])
            with pd.ExcelWriter(buf_extrato, engine="openpyxl") as writer:
                df_extrato.to_excel(writer, index=False, sheet_name="Extrato")
            st.download_button(
                f"📥 Exportar extrato de {cli['nome']} (Excel)", data=buf_extrato.getvalue(),
                file_name=f"extrato_{cli['nome'].lower().replace(' ', '_')}_{date.today().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"dl_extrato_{cli['id']}",
            )

            pill_map = {
                "ENTRADA":    ("Entrada", "#159A73", "#E1F6EE"),
                "UTILIZAÇÃO": ("Saída",   "#D6304A", "#FBE7EA"),
                "USO":        ("Saída",   "#D6304A", "#FBE7EA"),
                "ESTORNO":    ("Estorno", "#159A73", "#E1F6EE"),
                "AJUSTE":     ("Ajuste",  "#B4720A", "#FCEFD9"),
            }
            for l in linhas:
                label_p, cor_p, fundo_p = pill_map.get(l["tipo"], (l["tipo"], "#6E5A93", "#EDE9F8"))
                try:
                    data_fmt = pd.to_datetime(l["data"]).strftime("%d/%m/%Y")
                except Exception:
                    data_fmt = l["data"] or "—"
                sinal_txt  = "+" if l["valor"] >= 0 else ""
                cor_valor  = "#159A73" if l["valor"] >= 0 else "#D6304A"

                lcol1, lcol2, lcol3, lcol4 = st.columns([1, 5, 1.6, 0.5])
                lcol1.caption(data_fmt)
                with lcol2:
                    st.markdown(
                        f"<span style='background:{fundo_p};color:{cor_p};font-size:.68rem;"
                        f"font-weight:700;border-radius:20px;padding:2px 9px;margin-right:6px'>{label_p}</span>"
                        f"{l['desc']}", unsafe_allow_html=True)
                    if l["sub"]:
                        st.caption(l["sub"])
                lcol3.markdown(
                    f"<div style='text-align:right;font-family:monospace;color:{cor_valor}'>"
                    f"{sinal_txt}{brl(l['valor'])}</div>", unsafe_allow_html=True)
                if l["mov"] is not None:
                    if lcol4.button("🗑️", key=f"del_mov_{l['mov']['id']}", help="Apagar movimentação"):
                        _abrir_dialog_del_mov(l["mov"])
                st.markdown("<hr style='margin:4px 0;border-color:#EAE6F4'>", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 — RELATÓRIO MENSAL
# ══════════════════════════════════════════════════════════════════════════════
if main_tab == "📑 Relatório Mensal":
    from utils import MESES_PT

    hoje_r  = date.today()
    col_a, col_b = st.columns([1, 1])
    ano_r = col_a.selectbox("Ano", list(range(hoje_r.year, hoje_r.year - 4, -1)), key="rel_ano")
    mes_r = col_b.selectbox("Mês", list(range(1, 13)),
                             format_func=lambda m: MESES_PT[m],
                             index=hoje_r.month - 1, key="rel_mes")

    mes_ini = date(ano_r, mes_r, 1)
    mes_fim_day = (date(ano_r, mes_r % 12 + 1, 1) - timedelta(days=1)).day if mes_r < 12 else 31
    mes_fim = date(ano_r, mes_r, mes_fim_day)
    ini_s = mes_ini.strftime("%Y-%m-%d")
    fim_s = mes_fim.strftime("%Y-%m-%d")
    mes_label = f"{MESES_PT[mes_r]} {ano_r}"

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Dados do mês ─────────────────────────────────────────────────────────
    movs_mes = [m for m in movs_all if ini_s <= (m.get("data") or "")[:10] <= fim_s]
    notas_mes = [n for n in notas_all if ini_s <= (n.get("data_emissao") or "")[:10] <= fim_s]
    creds_novos = [c for c in creditos_all if any(n["id"] == c.get("nota_fiscal_id") for n in notas_mes)]

    total_consumido = sum(float(m.get("valor") or 0) for m in movs_mes
                          if m.get("tipo") in ("UTILIZAÇÃO", "USO"))
    total_novos = sum(float(c.get("valor_original") or 0) for c in creds_novos)

    # ── Saldo de cada crédito NO FIM do mês selecionado (não "hoje") ───────────
    # Créditos criados depois do fim do mês não contam; movimentações depois
    # do fim do mês também não — assim o relatório de um mês passado mostra a
    # posição REAL daquele momento, não o estado atual do sistema.
    _movs_por_credito = defaultdict(list)
    for m in movs_all:
        _movs_por_credito[m.get("credito_id")].append(m)

    def _saldo_credito_ate(cr, data_limite_s):
        criado = (cr.get("created_at") or "")[:10]
        if criado and criado > data_limite_s:
            return None  # crédito ainda não existia nessa data
        debito = 0.0
        for m in _movs_por_credito.get(cr["id"], []):
            d = (m.get("data") or "")[:10]
            if not d or d > data_limite_s:
                continue
            valor = m.get("valor") or 0
            debito += -valor if m.get("tipo") == "ESTORNO" else valor
        return (cr.get("valor_original") or 0) - debito

    _consumo_mes_por_cli = defaultdict(float)
    for m in movs_mes:
        if m.get("tipo") in ("UTILIZAÇÃO", "USO"):
            _consumo_mes_por_cli[m.get("cliente_nome", "")] += float(m.get("valor") or 0)

    _novos_mes_por_cli = defaultdict(float)
    for c in creds_novos:
        _novos_mes_por_cli[c.get("cliente_nome", "")] += float(c.get("valor_original") or 0)

    # ── Posição por cliente — saldo no fim do mês, mesmo sem movimento ─────────
    linhas_posicao = []
    saldo_fim_total = 0.0
    for cli in clientes_all:
        creds_cli = [c for c in creditos_all if c["cliente_id"] == cli["id"]]
        if not creds_cli:
            continue
        saldo_fim = sum((_saldo_credito_ate(c, fim_s) or 0) for c in creds_cli)
        movimentado = _consumo_mes_por_cli.get(cli["nome"], 0.0)
        novos = _novos_mes_por_cli.get(cli["nome"], 0.0)
        if abs(saldo_fim) < 0.005 and abs(movimentado) < 0.005 and abs(novos) < 0.005:
            continue  # nada a mostrar pra esse cliente nesse mês
        saldo_fim_total += saldo_fim
        linhas_posicao.append({
            "Cliente":               cli["nome"],
            "Saldo em fim do mês":   saldo_fim,
            "Movimentado no mês":    movimentado,
            "Créditos novos no mês": novos,
        })
    linhas_posicao.sort(key=lambda r: r["Saldo em fim do mês"], reverse=True)

    # ── KPIs ─────────────────────────────────────────────────────────────────
    def _kpi_card(col, label, valor, sub, cor):
        col.markdown(f"""
        <div style='background:#fff;border:1px solid #EAE6F4;border-radius:12px;
                    padding:14px 16px'>
          <div style='font-size:.68rem;text-transform:uppercase;letter-spacing:1.2px;
                      color:#8B6BAE;font-weight:700'>{label}</div>
          <div style='font-family:monospace;font-size:1.35rem;font-weight:700;
                      color:{cor};margin-top:3px'>{valor}</div>
          <div style='font-size:.74rem;color:#A899C4;margin-top:2px'>{sub}</div>
        </div>""", unsafe_allow_html=True)

    k1, k2, k3 = st.columns(3)
    _kpi_card(k1, "💸 Consumido no mês", brl(total_consumido),
              f"{len(movs_mes)} lançamento(s)", "#D6304A")
    _kpi_card(k2, "➕ Créditos novos", brl(total_novos),
              f"{len(creds_novos)} crédito(s)", "#159A73")
    _kpi_card(k3, f"💰 Saldo total em {mes_fim.strftime('%d/%m')}", brl(saldo_fim_total),
              f"{len(linhas_posicao)} cliente(s) com posição", "#7E16B8")

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Posição por cliente ──────────────────────────────────────────────────
    st.markdown(f"#### 👥 Posição por cliente — {mes_label}")
    if not linhas_posicao:
        st.info(f"Nenhum cliente com saldo, crédito novo ou consumo em {mes_label}.")
    else:
        df_pos = pd.DataFrame(linhas_posicao)
        df_pos_show = df_pos.copy()
        for _col_brl in ["Saldo em fim do mês", "Movimentado no mês", "Créditos novos no mês"]:
            df_pos_show[_col_brl] = df_pos_show[_col_brl].apply(brl)
        st.dataframe(df_pos_show, use_container_width=True, hide_index=True)

    # ── Detalhamento (fica escondido — só abre quem precisa) ───────────────────
    if movs_mes:
        with st.expander(f"📋 Detalhamento dos consumos de {mes_label} ({len(movs_mes)})"):
            df_rel = pd.DataFrame(movs_mes)
            df_rel["valor"] = pd.to_numeric(df_rel["valor"], errors="coerce").fillna(0)
            df_rel["data"]  = pd.to_datetime(df_rel["data"], errors="coerce")
            for col_opt in ["descricao_servico", "responsavel", "observacao"]:
                if col_opt not in df_rel.columns:
                    df_rel[col_opt] = None
            df_rel_show = df_rel.sort_values("data")[
                ["data", "cliente_nome", "descricao_servico", "valor", "responsavel", "observacao"]
            ].copy()
            df_rel_show["data"] = df_rel_show["data"].dt.strftime("%d/%m/%Y")
            df_rel_show.columns = ["Data", "Cliente", "Serviço", "Total", "Responsável", "Obs."]
            df_rel_show["Total"] = df_rel_show["Total"].apply(brl)
            st.dataframe(df_rel_show.fillna("—"), use_container_width=True, hide_index=True)

    # ── Geração do Excel ─────────────────────────────────────────────────────
    def _gerar_excel_relatorio(mes_label: str, movs: list, creds_new: list,
                                posicao: list, saldo_total: float) -> bytes:
        wb = Workbook()

        # Paleta
        PURPLE   = "4A1259"
        LAVENDER = "EDE9F8"
        WHITE    = "FFFFFF"
        GREY     = "F5F4FA"
        GREEN    = "D1FAE5"

        hdr_font  = Font(name="Arial", bold=True, color=WHITE, size=10)
        hdr_fill  = PatternFill("solid", fgColor=PURPLE)
        hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin      = Side(style="thin", color="CCCCCC")
        border    = Border(left=thin, right=thin, top=thin, bottom=thin)
        brl_fmt   = '#,##0.00'
        date_fmt  = 'DD/MM/YYYY'

        def _set_header(ws, cols, row=1):
            for c, (title, width) in enumerate(cols, 1):
                cell = ws.cell(row=row, column=c, value=title)
                cell.font    = hdr_font
                cell.fill    = hdr_fill
                cell.alignment = hdr_align
                cell.border  = border
                ws.column_dimensions[get_column_letter(c)].width = width
            ws.row_dimensions[row].height = 30

        def _style_data(ws, r, c, val, fmt=None, fill_color=None, bold=False):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font   = Font(name="Arial", size=9, bold=bold)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if fmt:
                cell.number_format = fmt
            if fill_color:
                cell.fill = PatternFill("solid", fgColor=fill_color)
            return cell

        # ── Aba 1: Resumo ────────────────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Resumo"

        ws1.merge_cells("A1:D1")
        t = ws1["A1"]
        t.value     = f"RELATÓRIO DE CRÉDITOS — {mes_label.upper()}"
        t.font      = Font(name="Arial", bold=True, color=WHITE, size=14)
        t.fill      = PatternFill("solid", fgColor=PURPLE)
        t.alignment = Alignment(horizontal="center", vertical="center")
        ws1.row_dimensions[1].height = 40

        ws1.merge_cells("A2:D2")
        ws1["A2"].value = f"Gerado em {date.today().strftime('%d/%m/%Y')} · Grupo GoGenetic"
        ws1["A2"].font  = Font(name="Arial", size=10, color="888888")
        ws1["A2"].alignment = Alignment(horizontal="center")
        ws1.row_dimensions[2].height = 20

        kpi_data = [
            ("Total consumido no mês",             total_consumido, brl_fmt, None),
            ("Novos créditos no mês",               total_novos,     brl_fmt, GREEN),
            (f"Saldo total em {mes_fim.strftime('%d/%m/%Y')}", saldo_total, brl_fmt, GREEN),
        ]
        ws1.row_dimensions[4].height = 20
        ws1.cell(4, 1, "INDICADORES DO MÊS").font = Font(name="Arial", bold=True, size=10, color=PURPLE)

        for i, (label, val, fmt, fc) in enumerate(kpi_data, 5):
            ws1.row_dimensions[i].height = 22
            c1 = ws1.cell(i, 1, label)
            c1.font   = Font(name="Arial", size=9, bold=True)
            c1.border = border
            c1.fill   = PatternFill("solid", fgColor=LAVENDER)
            ws1.column_dimensions["A"].width = 32
            c2 = ws1.cell(i, 2, val)
            c2.font   = Font(name="Arial", size=9)
            c2.border = border
            c2.number_format = fmt
            if fc:
                c2.fill = PatternFill("solid", fgColor=fc)
            ws1.column_dimensions["B"].width = 20

        # ── Aba 2: Posição por Cliente ───────────────────────────────────────
        ws2 = wb.create_sheet("Posição por Cliente")
        ws2.merge_cells("A1:D1")
        t2 = ws2["A1"]
        t2.value = f"POSIÇÃO POR CLIENTE — {mes_label.upper()}"
        t2.font  = Font(name="Arial", bold=True, color=WHITE, size=12)
        t2.fill  = PatternFill("solid", fgColor=PURPLE)
        t2.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[1].height = 35

        cols2 = [
            ("Cliente", 40), (f"Saldo em {mes_fim.strftime('%d/%m/%Y')} (R$)", 22),
            ("Movimentado no mês (R$)", 22), ("Créditos novos no mês (R$)", 22),
        ]
        _set_header(ws2, cols2, row=2)
        alt = False
        for r_idx, lin in enumerate(posicao, 3):
            alt = not alt
            fill_c = GREY if alt else WHITE
            ws2.row_dimensions[r_idx].height = 18
            _style_data(ws2, r_idx, 1, lin["Cliente"], fill_color=fill_c)
            _style_data(ws2, r_idx, 2, lin["Saldo em fim do mês"], brl_fmt, fill_c)
            _style_data(ws2, r_idx, 3, lin["Movimentado no mês"], brl_fmt, fill_c)
            _style_data(ws2, r_idx, 4, lin["Créditos novos no mês"], brl_fmt, fill_c)
        if posicao:
            tot_row = len(posicao) + 3
            ws2.row_dimensions[tot_row].height = 20
            _style_data(ws2, tot_row, 1, "TOTAL", fill_color=LAVENDER, bold=True)
            _style_data(ws2, tot_row, 2, saldo_total, brl_fmt, fill_color=LAVENDER, bold=True)
            _style_data(ws2, tot_row, 3, sum(l["Movimentado no mês"] for l in posicao),
                        brl_fmt, fill_color=LAVENDER, bold=True)
            _style_data(ws2, tot_row, 4, sum(l["Créditos novos no mês"] for l in posicao),
                        brl_fmt, fill_color=LAVENDER, bold=True)

        # ── Aba 3: Consumos Detalhados ───────────────────────────────────────
        ws3 = wb.create_sheet("Consumos Detalhados")
        ws3.merge_cells("A1:F1")
        t3 = ws3["A1"]
        t3.value = f"CONSUMOS DETALHADOS — {mes_label.upper()}"
        t3.font  = Font(name="Arial", bold=True, color=WHITE, size=12)
        t3.fill  = PatternFill("solid", fgColor=PURPLE)
        t3.alignment = Alignment(horizontal="center", vertical="center")
        ws3.row_dimensions[1].height = 35

        cols3 = [
            ("Data", 12), ("Cliente", 36), ("Serviço", 40),
            ("Total (R$)", 14), ("Responsável", 20), ("Observação", 30),
        ]
        _set_header(ws3, cols3, row=2)

        movs_sorted = sorted(movs, key=lambda m: (m.get("data") or ""))
        alt = False
        for r_idx, m in enumerate(movs_sorted, 3):
            alt = not alt
            fill_c = GREY if alt else WHITE
            ws3.row_dimensions[r_idx].height = 18
            data_val = None
            try:
                data_val = pd.to_datetime(m.get("data")).to_pydatetime() if m.get("data") else None
            except Exception:
                pass
            _style_data(ws3, r_idx, 1, data_val, date_fmt, fill_c)
            _style_data(ws3, r_idx, 2, m.get("cliente_nome", ""), fill_color=fill_c)
            _style_data(ws3, r_idx, 3, m.get("descricao_servico", "") or "—", fill_color=fill_c)
            _style_data(ws3, r_idx, 4, float(m.get("valor") or 0), brl_fmt, fill_c)
            _style_data(ws3, r_idx, 5, m.get("responsavel", "") or "—", fill_color=fill_c)
            _style_data(ws3, r_idx, 6, m.get("observacao", "") or "—", fill_color=fill_c)

        if movs_sorted:
            tot_row = len(movs_sorted) + 3
            ws3.row_dimensions[tot_row].height = 20
            ws3.merge_cells(f"A{tot_row}:C{tot_row}")
            tc = ws3[f"A{tot_row}"]
            tc.value  = "TOTAL DO MÊS"
            tc.font   = Font(name="Arial", bold=True, size=9)
            tc.fill   = PatternFill("solid", fgColor=LAVENDER)
            tc.border = border
            tv = ws3.cell(tot_row, 4)
            tv.value  = sum(float(m.get("valor") or 0) for m in movs_sorted)
            tv.number_format = brl_fmt
            tv.font   = Font(name="Arial", bold=True, size=9)
            tv.fill   = PatternFill("solid", fgColor=LAVENDER)
            tv.border = border

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ── Botão de download — gera Excel SOMENTE quando clicado ────────────────
    col_btn, _ = st.columns([2, 4])
    _rel_key = f"rel_{ano_r}_{mes_r:02d}"

    if col_btn.button(f"📥 Gerar Relatório Excel — {mes_label}",
                      use_container_width=True, key="btn_gerar_rel"):
        with st.spinner("Gerando Excel…"):
            st.session_state[_rel_key] = _gerar_excel_relatorio(
                mes_label, movs_mes, creds_novos, linhas_posicao, saldo_fim_total
            )

    if _rel_key in st.session_state:
        col_btn.download_button(
            label=f"⬇️ Baixar {mes_label}.xlsx",
            data=st.session_state[_rel_key],
            file_name=f"relatorio_creditos_{ano_r}_{mes_r:02d}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key="btn_dl_rel",
        )
